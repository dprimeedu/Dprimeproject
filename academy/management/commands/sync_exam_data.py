"""
모의고사 엑셀 → SQLite 동기화

처리하는 시트 (Phase 1 완성판):
    문제출력  → question_data (단독 테이블, KEY_TABLE 무관)
    그 외 14개 → 각 detail 테이블, KEY_TABLE 경유

사용:
    python manage.py sync_exam_data --dry-run            # 변경 미리보기
    python manage.py sync_exam_data                       # 실제 적용
    python manage.py sync_exam_data --grade 고1           # 한 학년만
    python manage.py sync_exam_data --sheet 내신단어      # 특정 시트만
    python manage.py sync_exam_data --limit 50            # 시트당 50행만

정책:
    - 자연키로 행 식별 (시트마다 다름 — SHEET_SPECS 참조)
    - 추가/수정만 — 삭제는 안 함
    - 자동 ID 발급 (Index, PK_number)
"""
import os
import time
import traceback
from pathlib import Path

from django.core.management.base import BaseCommand
from django.db import connection, transaction
from django.db.models import Max
from django.utils import timezone

import openpyxl

from academy.models import QuestionData, SyncLog


DEFAULT_EXCEL_DIR = os.environ.get(
    "EXAM_EXCEL_DIR",
    r"Z:\home\Drive\문서 자동화 작업\엑셀자료모음\모의고사 EXCEL",
)

GRADE_FILES = {
    "고1": "고1 전체 모의고사 DB.xlsx",
    "고2": "고2 전체 모의고사 DB.xlsx",
    "고3": "고3 전체 모의고사 DB.xlsx",
}

# Detail 테이블 시트 명세
# cols: {DB컬럼: 엑셀컬럼}
# natural_key: 같은 PK_number 안에서 행을 구별하는 컬럼 (빈 리스트면 1:1 — PK_number만으로 식별)
# year_col: 연도 컬럼 이름 (기본 '연도', 일부 시트는 '년')
SHEET_SPECS = {
    "내신단어": {
        "table": "WordTest",
        "cols": {
            "word": "단어",
            "korean_definition": "해석",
        },
        "optional_cols": {
            # EBS '내신단어'에는 '영영' 없음 (별도 '영영단어' 시트 존재)
            "english_definition": "영영",
        },
        "natural_key": ["word"],
    },
    "상세해설": {
        "table": "DetailedExplanation",
        "cols": {"Saved_location": "위치"},
        "natural_key": [],
    },
    "어법1단계": {
        "table": "Grammarlv1",
        "cols": {"Question": "원문", "answer": "정답"},
        "natural_key": [],
    },
    "어법2단계": {
        "table": "Grammarlv2",
        "cols": {"Question": "문제", "Answer": "정답"},
        "natural_key": ["Question"],
    },
    "어법3단계": {
        "table": "Grammarlv3",
        "cols": {"Question": "문제", "Answer": "정답"},
        "natural_key": ["Question"],
    },
    "중요영작": {
        "table": "Translation",
        "cols": {
            # EBS는 컬럼명이 다름: 영어/한글, '회화..'/'중요영작' 컬럼은 부재
            "Sentence": ["영작", "영어"],
            "Translation": ["해석", "한글"],
        },
        "optional_cols": {
            "ETC": "회화/Extra/기타여부",
            "Key_sentence": "중요영작",
        },
        "natural_key": ["Sentence"],
    },
    "요약문완성": {
        "table": "Summary",
        "cols": {
            "Origin_text": "원문",
            "Red": "빨",
            "Blue": "파",
            "summary": "요약문",
            "Answer": "정답",
        },
        "natural_key": [],
    },
    "원문추가": {
        "table": "Additional_text",
        "cols": {"Additional_text": "원문추가"},
        "natural_key": [],
        "year_col": "년",
    },
    "내신TEST": {
        "table": "SchoolExamtest",
        "cols": {
            "Question": "문제",
            "Type": "유형",
            "Sentence": "지문",
            "Option": "보기",
            "Answer": "정답",
            "Modified": "변형",
        },
        "natural_key": ["Question"],
    },
    "서술형": {
        "table": "Descriptive_Question",
        "cols": {"Que_Location": "위치", "Ans_Location": "정답"},
        "natural_key": [],
    },
    "내신빨파": {
        "table": "RedBlue",
        "cols": {
            "Origin_text": "원문",
            "Red": "빨",
            "Blue": "파",
            "Ans_location": "정답",
        },
        "natural_key": [],
    },
    "객관식빈칸": {
        "table": "FillinBlank",
        "cols": {
            "Question": "문제",
            "Options": "보기",
            "Answer": "정답",
        },
        "optional_cols": {
            # EBS 객관식빈칸은 '지문' 없이 보기만으로 출제하는 형태가 있음
            "Sentence": "지문",
        },
        "natural_key": ["Question"],
    },
    "원문": {
        "table": "Original_text",
        "cols": {"Origin_text": ["원문", "지문"]},  # 고2는 '지문'으로 돼있음
        "natural_key": [],
    },
    "보기변형": {
        "table": "Modified_Questions",
        "cols": {
            "Question": "문제",
            "Qtype": "유형",
            "Sentence": "지문",
            "Option": "보기",
            "Answer": "정답",
            # EBS/교과서 통합본은 '변형', 모의고사 통합본은 '(변형)정답'
            "Modified": ["(변형)정답", "변형"],
        },
        "natural_key": ["Question"],
    },
}

# 시트명 alias — 통합본마다 시트 제목이 다르면 spec 키로 매핑
SHEET_NAME_ALIASES = {
    "변형문제": "보기변형",   # EBS/교과서 통합본
}

# 처리 순서: 문제출력 먼저, 나머지는 SHEET_SPECS 순서대로
SHEETS_ORDERED = ["문제출력"] + list(SHEET_SPECS.keys())


def _s(v) -> str:
    """엑셀 셀 값 → 안전 문자열 (None → '')"""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _to_int(v):
    """엑셀 값 → int (실패 시 None)"""
    if v is None:
        return None
    try:
        if isinstance(v, str):
            v = v.strip()
            if not v:
                return None
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _resolve_col(col_dict, name_or_names):
    """단일 이름 또는 alias 리스트 → 첫 매칭 인덱스 또는 None"""
    if isinstance(name_or_names, str):
        return col_dict.get(name_or_names)
    for n in name_or_names:
        if n in col_dict:
            return col_dict[n]
    return None


def _parse_unit_ymn(unit_str):
    """단원 문자열 '2019-03-18' → (2019, 3, 18). 실패 시 (None, None, None)"""
    if not unit_str:
        return None, None, None
    s = _s(unit_str)
    parts = s.split("-")
    if len(parts) != 3:
        return None, None, None
    try:
        return int(parts[0]), int(parts[1]), int(parts[2])
    except ValueError:
        return None, None, None


class Command(BaseCommand):
    help = "엑셀 모의고사 DB → SQLite 동기화 (15개 시트 전체)"

    def add_arguments(self, parser):
        parser.add_argument("--dry-run", action="store_true", help="변경 미리보기만")
        parser.add_argument(
            "--grade",
            choices=["고1", "고2", "고3", "all"],
            default="all",
        )
        parser.add_argument(
            "--sheet",
            choices=SHEETS_ORDERED + ["all"],
            default="all",
            help="특정 시트만 처리",
        )
        parser.add_argument("--path", default=DEFAULT_EXCEL_DIR)
        parser.add_argument("--limit", type=int, default=None, help="시트당 행 수 제한")
        parser.add_argument(
            "--triggered-by",
            choices=["cli", "web", "auto"],
            default="cli",
            help="누가 트리거했는지 (SyncLog 기록용)",
        )
        parser.add_argument(
            "--no-log",
            action="store_true",
            help="SyncLog 기록 비활성화 (디버깅 시)",
        )

    def handle(self, *args, **options):
        self.dry_run = options["dry_run"]
        self.limit = options["limit"]
        base_path = Path(options["path"])
        log_enabled = not options["no_log"]

        if self.dry_run:
            self.stdout.write(self.style.WARNING("🔸 DRY RUN — 실제 DB는 변경되지 않습니다\n"))

        if not base_path.exists():
            self.stderr.write(self.style.ERROR(f"❌ 경로 없음: {base_path}"))
            return

        grades = ["고1", "고2", "고3"] if options["grade"] == "all" else [options["grade"]]
        sheets = SHEETS_ORDERED if options["sheet"] == "all" else [options["sheet"]]

        total = {"added": 0, "updated": 0, "skipped": 0, "error": 0}
        sheet_results = {}  # {grade: {sheet: {added, updated, ...}}}
        self.error_details = []  # 누적

        # SyncLog 시작 기록
        sync_log = None
        if log_enabled:
            sync_log = SyncLog.objects.create(
                triggered_by=options["triggered_by"],
                dry_run=self.dry_run,
                target_grade=options["grade"],
                target_sheet=options["sheet"],
            )

        started_at = time.time()

        # 트랜잭션으로 묶어 성능 향상 (실패 시 rollback)
        outer = transaction.atomic() if not self.dry_run else _NoopCtx()

        with outer:
            for grade in grades:
                file_path = base_path / GRADE_FILES[grade]
                if not file_path.exists():
                    self.stderr.write(self.style.WARNING(f"⚠️  파일 없음: {file_path.name}"))
                    continue

                self.stdout.write(f"\n📊 [{grade}] {file_path.name}")

                try:
                    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                except Exception as e:
                    self.stderr.write(self.style.ERROR(f"❌ 파일 열기 실패: {e}"))
                    continue

                for sheet_name in sheets:
                    if sheet_name not in wb.sheetnames:
                        continue

                    ws = wb[sheet_name]

                    if sheet_name == "문제출력":
                        stats = self._sync_munje_chulryeok(ws, grade)
                    else:
                        spec = SHEET_SPECS[sheet_name]
                        stats = self._sync_legacy_table(ws, grade, sheet_name, spec)

                    line = (
                        f"  [{sheet_name:<8}] "
                        f"추가 {stats['added']:>5}, 수정 {stats['updated']:>5}, "
                        f"변경없음 {stats['skipped']:>5}"
                    )
                    if stats["error"]:
                        line += f", 오류 {stats['error']}"
                    self.stdout.write(self.style.SUCCESS(line))

                    for k in total:
                        total[k] += stats[k]

                    sheet_results.setdefault(grade, {})[sheet_name] = stats

                wb.close()

        duration = time.time() - started_at

        self.stdout.write("")
        self.stdout.write(
            self.style.SUCCESS(
                f"{'(dry-run) ' if self.dry_run else ''}"
                f"✅ 완료: 추가 {total['added']}, 수정 {total['updated']}, "
                f"변경없음 {total['skipped']}, 오류 {total['error']} "
                f"({duration:.1f}초)"
            )
        )

        # SyncLog 완료 기록
        if sync_log is not None:
            sync_log.finished_at = timezone.now()
            sync_log.duration_seconds = duration
            sync_log.added = total["added"]
            sync_log.updated = total["updated"]
            sync_log.skipped = total["skipped"]
            sync_log.errors = total["error"]
            sync_log.sheet_results = sheet_results
            sync_log.error_details = self.error_details
            sync_log.save()
            self.stdout.write(
                f"   📋 SyncLog #{sync_log.id} 기록됨 — /admin/academy/synclog/{sync_log.id}/"
            )

    # ───────────────────────────────────────────────────────────
    # 헬퍼
    # ───────────────────────────────────────────────────────────

    def _get_or_create_key(self, cursor, grade, year, month, number, qtype=""):
        """KEY_TABLE에서 (grade, year, month, number, Qtype) 조회/생성, PK_number 반환"""
        cursor.execute(
            'SELECT PK_number FROM KEY_TABLE '
            "WHERE grade=? AND year=? AND month=? AND number=? AND Qtype=?",
            (grade, year, month, number, qtype),
        )
        row = cursor.fetchone()
        if row:
            return row[0]

        if self.dry_run:
            # dry-run에선 가짜 PK로 동작 (실제 INSERT 안 함)
            return -1

        cursor.execute('SELECT COALESCE(MAX(PK_number), 0) FROM KEY_TABLE')
        next_pk = cursor.fetchone()[0] + 1
        total_number = f"{year}-{month:02d}-{number:02d}"
        cursor.execute(
            'INSERT INTO KEY_TABLE (PK_number, Total_number, grade, year, month, number, Qtype) '
            "VALUES (?, ?, ?, ?, ?, ?, ?)",
            (next_pk, total_number, grade, year, month, number, qtype),
        )
        return next_pk

    # ───────────────────────────────────────────────────────────
    # 문제출력 → question_data (단독)
    # ───────────────────────────────────────────────────────────

    def _sync_munje_chulryeok(self, ws, grade):
        stats = {"added": 0, "updated": 0, "skipped": 0, "error": 0}

        rows = ws.iter_rows(values_only=True)
        try:
            headers = list(next(rows))
        except StopIteration:
            return stats

        col = {h: i for i, h in enumerate(headers) if h}

        required = ["문제", "유형", "연도", "월", "번호"]
        missing = [c for c in required if c not in col]
        if missing:
            self.stderr.write(self.style.ERROR(f"    필수 컬럼 누락: {missing}"))
            return stats

        try:
            max_idx_raw = QuestionData.objects.aggregate(Max("색인"))["색인__max"]
            next_idx = int(max_idx_raw) + 1 if max_idx_raw else 1
        except (TypeError, ValueError):
            next_idx = 1

        count = 0
        for row in rows:
            if self.limit and count >= self.limit:
                break
            count += 1

            if not any(c is not None and str(c).strip() for c in row):
                continue

            try:
                연도 = _s(row[col["연도"]])
                강 = _s(row[col["월"]])
                번호_raw = row[col["번호"]]
                유형 = _s(row[col["유형"]])

                if not 연도 or not 강 or 번호_raw is None:
                    stats["skipped"] += 1
                    continue

                번호 = int(번호_raw)

                data = {
                    "문제": _s(row[col["문제"]]),
                    "유형": 유형,
                    "지문": _s(row[col["지문"]]) if "지문" in col else "",
                    "보기": _s(row[col["보기"]]) if "보기" in col else "",
                    "정답": _s(row[col["정답"]]) if "정답" in col else "",
                    "변형": "",
                    "학년": grade,
                    "연도": 연도,
                    "강": 강,
                    "번호": 번호,
                    "단원": _s(row[col["단원"]]) if "단원" in col else "",
                    "그림": _s(row[col["그림"]]) if "그림" in col else "",
                }

                existing = QuestionData.objects.filter(
                    학년=grade, 연도=연도, 강=강, 번호=번호, 유형=유형
                ).first()

                if existing:
                    changed = []
                    for k, v in data.items():
                        if k in ("학년", "연도", "강", "번호", "유형"):
                            continue
                        if _s(getattr(existing, k)) != _s(v):
                            changed.append(k)

                    if changed:
                        if not self.dry_run:
                            for k in changed:
                                setattr(existing, k, data[k])
                            existing.save()
                        stats["updated"] += 1
                    else:
                        stats["skipped"] += 1
                else:
                    if not self.dry_run:
                        data["색인"] = str(next_idx)
                        next_idx += 1
                        QuestionData.objects.create(**data)
                    else:
                        next_idx += 1
                    stats["added"] += 1

            except Exception as e:
                stats["error"] += 1
                self.error_details.append({
                    "sheet": getattr(ws, "title", "?"),
                    "row": count + 1,
                    "message": str(e),
                })
                if stats["error"] <= 3:
                    self.stderr.write(self.style.WARNING(f"    행 {count} 오류: {e}"))

        return stats

    # ───────────────────────────────────────────────────────────
    # 일반 detail 시트 (KEY_TABLE 경유)
    # ───────────────────────────────────────────────────────────

    def _sync_legacy_table(self, ws, grade, sheet_name, spec):
        stats = {"added": 0, "updated": 0, "skipped": 0, "error": 0}

        rows = ws.iter_rows(values_only=True)
        try:
            headers = list(next(rows))
        except StopIteration:
            return stats

        col = {h: i for i, h in enumerate(headers) if h}

        # 연도/월/번호 컬럼 인덱스 (없으면 None, 단원 파싱으로 폴백)
        year_idx = _resolve_col(col, [spec.get("year_col", "연도"), "연도", "년"])
        month_idx = col.get("월")
        number_idx = col.get("번호")
        unit_idx = col.get("단원")

        # 셋 다 없고 단원 파싱도 불가능하면 스킵
        if (year_idx is None or month_idx is None or number_idx is None) and unit_idx is None:
            self.stderr.write(
                self.style.WARNING(
                    f"    {sheet_name}: 연도/월/번호 + 단원 모두 누락 — 스킵"
                )
            )
            return stats

        # DB 컬럼 → 엑셀 인덱스 (alias 지원: 값이 리스트면 그 중 하나 매칭)
        excel_idx = {}
        for db_col, excel_col_spec in spec["cols"].items():
            idx = _resolve_col(col, excel_col_spec)
            if idx is None:
                names = excel_col_spec if isinstance(excel_col_spec, list) else [excel_col_spec]
                self.stderr.write(
                    self.style.WARNING(
                        f"    {sheet_name}: 엑셀 컬럼 {names} 없음 (DB {db_col})"
                    )
                )
                return stats
            excel_idx[db_col] = idx

        # optional 컬럼 — 없으면 빈 문자열로 채움
        optional_idx = {}
        for db_col, excel_col_spec in spec.get("optional_cols", {}).items():
            idx = _resolve_col(col, excel_col_spec)
            if idx is not None:
                optional_idx[db_col] = idx

        table = spec["table"]
        qtype = spec.get("qtype", "")
        natural_key = spec.get("natural_key", [])

        cursor = connection.cursor()

        # Index 자동 발급
        cursor.execute(f'SELECT COALESCE(MAX("Index"), 0) FROM "{table}"')
        next_index = cursor.fetchone()[0] + 1

        count = 0
        for row in rows:
            if self.limit and count >= self.limit:
                break
            count += 1

            if not any(c is not None and str(c).strip() for c in row):
                continue

            try:
                year = _to_int(row[year_idx]) if year_idx is not None else None
                month = _to_int(row[month_idx]) if month_idx is not None else None
                number = _to_int(row[number_idx]) if number_idx is not None else None

                # 단원 컬럼에서 폴백 파싱 (예: '2019-03-18')
                if (year is None or month is None or number is None) and unit_idx is not None:
                    py, pm, pn = _parse_unit_ymn(row[unit_idx])
                    year = year if year is not None else py
                    month = month if month is not None else pm
                    number = number if number is not None else pn

                if year is None or month is None or number is None:
                    stats["skipped"] += 1
                    continue

                # 데이터 추출 (필수 + optional, optional은 부재시 빈 문자열)
                data = {db_col: _s(row[idx]) for db_col, idx in excel_idx.items()}
                for db_col in spec.get("optional_cols", {}):
                    data[db_col] = _s(row[optional_idx[db_col]]) if db_col in optional_idx else ""

                # KEY_TABLE 조회/생성
                pk = self._get_or_create_key(cursor, grade, year, month, number, qtype)

                # natural key 로 detail 행 조회
                if natural_key:
                    conditions = ["PK_number = ?"] + [f'"{c}" = ?' for c in natural_key]
                    params = [pk] + [data[c] for c in natural_key]
                else:
                    conditions = ["PK_number = ?"]
                    params = [pk]
                where = " AND ".join(conditions)

                cursor.execute(f'SELECT "Index" FROM "{table}" WHERE {where}', params)
                existing = cursor.fetchone()

                if existing:
                    idx_val = existing[0]
                    # 변경 검사
                    select_cols = ", ".join(f'"{c}"' for c in data.keys())
                    cursor.execute(
                        f'SELECT {select_cols} FROM "{table}" WHERE "Index" = ?', [idx_val]
                    )
                    current = cursor.fetchone()
                    changed = []
                    for i, (db_col, new_val) in enumerate(data.items()):
                        if _s(current[i]) != _s(new_val):
                            changed.append(db_col)

                    if changed:
                        if not self.dry_run:
                            set_clause = ", ".join(f'"{c}" = ?' for c in changed)
                            update_params = [data[c] for c in changed] + [idx_val]
                            cursor.execute(
                                f'UPDATE "{table}" SET {set_clause} WHERE "Index" = ?',
                                update_params,
                            )
                        stats["updated"] += 1
                    else:
                        stats["skipped"] += 1
                else:
                    if not self.dry_run:
                        cols_list = ['"Index"', '"PK_number"'] + [
                            f'"{c}"' for c in data.keys()
                        ]
                        placeholders = ", ".join("?" * (len(data) + 2))
                        values = [next_index, pk] + list(data.values())
                        cursor.execute(
                            f'INSERT INTO "{table}" ({", ".join(cols_list)}) '
                            f"VALUES ({placeholders})",
                            values,
                        )
                    next_index += 1
                    stats["added"] += 1

            except Exception as e:
                stats["error"] += 1
                if stats["error"] <= 3:
                    self.stderr.write(
                        self.style.WARNING(f"    [{sheet_name}] 행 {count} 오류: {e}")
                    )

        return stats


class _NoopCtx:
    """dry-run 시 transaction 미사용"""
    def __enter__(self):
        return self
    def __exit__(self, *args):
        return False
