"""
EBS/교과서 통합본 → SQLite 동기화

폴더 walk로 *통합*.xlsx 를 찾아서, sync_exam_data 와 같은 detail 테이블에 적재한다.
모의고사와 다른 점:
    - 단원 형식이 '1-1', '1-A' 같은 자유 텍스트
    - KEY_TABLE 자연키 = (category, book, total_number)
    - year/month/number 는 단원 파싱이 가능하면 채우고, 아니면 ''

사용:
    python manage.py sync_textbook_data --category ebs_high3 --dry-run
    python manage.py sync_textbook_data --category ebs_naesin
    python manage.py sync_textbook_data --category textbook --limit-files 3
    python manage.py sync_textbook_data --category all
"""
import os
import re
import time
import traceback
from pathlib import Path

from django.core.management.base import BaseCommand
from django.db import connection, transaction
from django.utils import timezone

import openpyxl

from academy.models import SyncLog
from academy.management.commands.sync_exam_data import (
    SHEET_SPECS,
    SHEET_NAME_ALIASES,
    _s,
    _resolve_col,
)


CATEGORY_PATHS = {
    'ebs_high3':  r'Z:\home\Drive\교재폴더\EBS\1. 고3 연계교재',
    'ebs_naesin': r'Z:\home\Drive\교재폴더\EBS\2. 내신관련 교재',
    'textbook':   r'Z:\home\Drive\교재폴더\내신',
}

# 카테고리 → 기본 학년 (파일명/경로에서 인식 실패 시 폴백)
CATEGORY_DEFAULT_GRADE = {
    'ebs_high3':  '고3',
    'ebs_naesin': '',   # 파일명/경로에서 추출
    'textbook':   '',   # 〃
}

# 교과서/EBS detail 시트들 (KEY_TABLE 시트는 제외 — 그건 통합본 내부용)
TEXTBOOK_SHEETS = [
    s for s in SHEET_SPECS.keys()
    if s not in ('어법1단계',)  # 어법1단계는 Question 컬럼명을 자동 매칭하므로 OK
] + ['어법1단계']

GRADE_RE = re.compile(r'(고\s*[1-3]|중\s*[1-3])')


class _NoopCtx:
    def __enter__(self):
        return self
    def __exit__(self, *args):
        return False


def _extract_grade(file_path: Path, fallback: str = ''):
    """파일명 / 경로에서 학년 추출 (고1/고2/고3/중1/중2/중3)."""
    text = str(file_path)
    m = GRADE_RE.search(text)
    if m:
        return m.group(1).replace(' ', '')
    return fallback


def _extract_book(file_path: Path) -> str:
    """파일명에서 '통합' 접미사 제거 → book 식별자."""
    name = file_path.stem  # ex: '2024년 수능완성 통합'
    name = re.sub(r'\s*\(.*?\)\s*$', '', name)  # 끝에 '(수정전)' 같은 괄호 제거
    name = re.sub(r'\s*통합\s*$', '', name).strip()
    return name


def _to_int_or_none(v):
    if v is None:
        return None
    try:
        if isinstance(v, str):
            v = v.strip()
            if not v:
                return None
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _parse_unit_token(unit_str: str):
    """
    '1-1' / '1-A' / '12-3' 형식 → (단원, 번호) 문자열 튜플.
    매칭 실패 시 (전체, '').
    """
    s = _s(unit_str)
    if not s:
        return '', ''
    m = re.match(r'^(\S+?)\s*[-–]\s*(\S+)$', s)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return s, ''


def _walk_books(category_path: Path, skip_keywords=()):
    """카테고리 폴더 아래에서 *통합*.xlsx 파일을 yield. 락파일/백업 제외."""
    for root, _, files in os.walk(category_path):
        for f in files:
            if not f.lower().endswith('.xlsx'):
                continue
            if f.startswith('~$'):
                continue
            if '통합' not in f:
                continue
            # 백업/복사본 등 제외
            low = f.lower()
            if any(kw in low for kw in ('백업', '복사본', '수정전', '(원본)')):
                continue
            if any(kw in f for kw in skip_keywords):
                continue
            yield Path(root) / f


class Command(BaseCommand):
    help = "EBS/교과서 통합본 → SQLite 동기화 (KeyTable + 13 detail 테이블)"

    def add_arguments(self, parser):
        parser.add_argument(
            '--category',
            choices=list(CATEGORY_PATHS.keys()) + ['all'],
            required=True,
            help='동기화할 카테고리',
        )
        parser.add_argument('--dry-run', action='store_true', help='변경 미리보기만')
        parser.add_argument(
            '--limit-files', type=int, default=None,
            help='카테고리당 최대 파일 수 (테스트용)',
        )
        parser.add_argument(
            '--file', default=None,
            help='특정 파일 경로 하나만 처리 (테스트용)',
        )
        parser.add_argument(
            '--year', default=None,
            help='EBS 고3 카테고리에서 특정 연도 하위만 처리 (e.g., "2024년")',
        )
        parser.add_argument(
            '--limit', type=int, default=None,
            help='시트당 행 수 제한 (테스트용)',
        )
        parser.add_argument(
            '--triggered-by',
            choices=['cli', 'web', 'auto'],
            default='cli',
        )
        parser.add_argument(
            '-v2', '--verbose-skip', action='store_true', dest='verbose_skip',
            help='시트 스킵 사유 출력',
        )

    def handle(self, *args, **options):
        self.dry_run = options['dry_run']
        self.limit = options['limit']
        self.verbose = options.get('verbose_skip', False)
        cat_opt = options['category']
        limit_files = options['limit_files']

        if self.dry_run:
            self.stdout.write(self.style.WARNING(
                '🔸 DRY RUN — 실제 DB는 변경되지 않습니다\n'))

        single_file = options.get('file')
        year_filter = options.get('year')

        cats = list(CATEGORY_PATHS.keys()) if cat_opt == 'all' else [cat_opt]

        total = {'added': 0, 'updated': 0, 'skipped': 0, 'error': 0}
        sheet_results = {}
        self.error_details = []

        sync_log = SyncLog.objects.create(
            triggered_by=options['triggered_by'],
            dry_run=self.dry_run,
            target_grade=cat_opt,
            target_sheet='textbook',
        )

        started_at = time.time()
        outer = transaction.atomic() if not self.dry_run else _NoopCtx()

        with outer:
            for cat in cats:
                base = Path(CATEGORY_PATHS[cat])
                if year_filter:
                    base = base / year_filter
                if not base.exists():
                    self.stderr.write(self.style.ERROR(f'❌ 경로 없음: {base}'))
                    continue

                if single_file:
                    files = [Path(single_file)]
                else:
                    files = list(_walk_books(base))
                if limit_files:
                    files = files[:limit_files]

                self.stdout.write(self.style.SUCCESS(
                    f'\n=== [{cat}] {len(files)}개 통합본 ==='))

                for fp in files:
                    book = _extract_book(fp)
                    grade = _extract_grade(fp, CATEGORY_DEFAULT_GRADE.get(cat, ''))
                    self.stdout.write(f'\n📘 [{book}] (grade={grade or "?"}) {fp.name}')

                    try:
                        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
                    except Exception as e:
                        self.stderr.write(self.style.ERROR(f'❌ 열기 실패: {e}'))
                        total['error'] += 1
                        self.error_details.append({'file': str(fp), 'message': str(e)})
                        continue

                    file_stats = {'added': 0, 'updated': 0, 'skipped': 0, 'error': 0}

                    # 엑셀에 실제 있는 시트들을 spec 키(또는 alias)로 매핑
                    sheet_to_spec = {}
                    for ws_name in wb.sheetnames:
                        canon = SHEET_NAME_ALIASES.get(ws_name, ws_name)
                        if canon in SHEET_SPECS:
                            sheet_to_spec[ws_name] = canon

                    for ws_name, spec_key in sheet_to_spec.items():
                        ws = wb[ws_name]
                        spec = SHEET_SPECS[spec_key]
                        sheet_name = spec_key  # 출력/통계용
                        try:
                            stats = self._sync_sheet(
                                ws, cat, book, grade, sheet_name, spec)
                        except Exception as e:
                            self.stderr.write(self.style.ERROR(
                                f'  [{sheet_name}] 처리 실패: {e}'))
                            stats = {'added': 0, 'updated': 0, 'skipped': 0, 'error': 1}
                            self.error_details.append({
                                'file': str(fp), 'sheet': sheet_name,
                                'message': str(e), 'trace': traceback.format_exc()[:500],
                            })

                        for k in file_stats:
                            file_stats[k] += stats[k]
                            total[k] += stats[k]

                        if stats['added'] or stats['updated']:
                            self.stdout.write(
                                f'  [{sheet_name:<8}] +{stats["added"]:>4} '
                                f'~{stats["updated"]:>4} ={stats["skipped"]:>4}'
                                + (f' !{stats["error"]}' if stats['error'] else '')
                            )

                    wb.close()
                    sheet_results.setdefault(cat, {})[book] = file_stats

        duration = time.time() - started_at

        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS(
            f'{"(dry-run) " if self.dry_run else ""}✅ 완료: '
            f'+{total["added"]} ~{total["updated"]} ={total["skipped"]} '
            f'!{total["error"]} ({duration:.1f}초)'
        ))

        sync_log.finished_at = timezone.now()
        sync_log.duration_seconds = duration
        sync_log.added = total['added']
        sync_log.updated = total['updated']
        sync_log.skipped = total['skipped']
        sync_log.errors = total['error']
        sync_log.sheet_results = sheet_results
        sync_log.error_details = self.error_details
        sync_log.save()
        self.stdout.write(f'   📋 SyncLog #{sync_log.id} 기록됨')

    # ───────────────────────────────────────────────────────────

    def _get_or_create_key(self, cursor, category, book, grade, total_number, qtype=''):
        """
        (category, book, Total_number) 자연키로 KEY_TABLE 조회/생성.
        반환: PK_number
        """
        cursor.execute(
            'SELECT PK_number FROM KEY_TABLE '
            'WHERE category=? AND book=? AND Total_number=? AND Qtype=?',
            (category, book, total_number, qtype),
        )
        row = cursor.fetchone()
        if row:
            return row[0]

        if self.dry_run:
            return -1

        cursor.execute('SELECT COALESCE(MAX(PK_number), 0) FROM KEY_TABLE')
        next_pk = cursor.fetchone()[0] + 1

        unit, num = _parse_unit_token(total_number)
        cursor.execute(
            'INSERT INTO KEY_TABLE '
            '(PK_number, Total_number, grade, year, month, number, Qtype, category, book) '
            'VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)',
            (next_pk, total_number, grade or '', '', unit, num, qtype, category, book),
        )
        return next_pk

    def _sync_sheet(self, ws, category, book, grade, sheet_name, spec):
        stats = {'added': 0, 'updated': 0, 'skipped': 0, 'error': 0}

        rows = ws.iter_rows(values_only=True)
        try:
            headers = list(next(rows))
        except StopIteration:
            return stats

        col = {h: i for i, h in enumerate(headers) if h}

        unit_idx = col.get('단원')
        if unit_idx is None:
            if self.verbose:
                self.stdout.write(self.style.WARNING(
                    f'    [{sheet_name}] 단원 컬럼 없음 — 스킵 (headers={list(col.keys())[:6]})'))
            return stats

        excel_idx = {}
        for db_col, excel_col_spec in spec['cols'].items():
            idx = _resolve_col(col, excel_col_spec)
            if idx is None:
                if self.verbose:
                    names = excel_col_spec if isinstance(excel_col_spec, list) else [excel_col_spec]
                    self.stdout.write(self.style.WARNING(
                        f'    [{sheet_name}] 컬럼 {names} 없음 — 스킵 (headers={list(col.keys())[:6]})'))
                return stats
            excel_idx[db_col] = idx

        optional_idx = {}
        for db_col, excel_col_spec in spec.get('optional_cols', {}).items():
            idx = _resolve_col(col, excel_col_spec)
            if idx is not None:
                optional_idx[db_col] = idx

        table = spec['table']
        qtype = spec.get('qtype', '')
        natural_key = spec.get('natural_key', [])

        cursor = connection.cursor()
        cursor.execute(f'SELECT COALESCE(MAX("Index"), 0) FROM "{table}"')
        next_index = cursor.fetchone()[0] + 1

        count = 0
        for row in rows:
            if self.limit and count >= self.limit:
                break
            count += 1

            if not any(c is not None and str(c).strip() for c in row):
                continue

            try:
                total_number = _s(row[unit_idx])
                if not total_number:
                    stats['skipped'] += 1
                    continue

                data = {db_col: _s(row[idx]) for db_col, idx in excel_idx.items()}
                for db_col in spec.get('optional_cols', {}):
                    data[db_col] = _s(row[optional_idx[db_col]]) if db_col in optional_idx else ''

                pk = self._get_or_create_key(
                    cursor, category, book, grade, total_number, qtype)

                if natural_key:
                    conditions = ['PK_number = ?'] + [f'"{c}" = ?' for c in natural_key]
                    params = [pk] + [data[c] for c in natural_key]
                else:
                    conditions = ['PK_number = ?']
                    params = [pk]
                where = ' AND '.join(conditions)

                cursor.execute(f'SELECT "Index" FROM "{table}" WHERE {where}', params)
                existing = cursor.fetchone()

                if existing:
                    idx_val = existing[0]
                    select_cols = ', '.join(f'"{c}"' for c in data.keys())
                    cursor.execute(
                        f'SELECT {select_cols} FROM "{table}" WHERE "Index" = ?', [idx_val]
                    )
                    current = cursor.fetchone()
                    changed = []
                    for i, (db_col, new_val) in enumerate(data.items()):
                        if _s(current[i]) != _s(new_val):
                            changed.append(db_col)

                    if changed:
                        if not self.dry_run:
                            set_clause = ', '.join(f'"{c}" = ?' for c in changed)
                            update_params = [data[c] for c in changed] + [idx_val]
                            cursor.execute(
                                f'UPDATE "{table}" SET {set_clause} WHERE "Index" = ?',
                                update_params,
                            )
                        stats['updated'] += 1
                    else:
                        stats['skipped'] += 1
                else:
                    if not self.dry_run:
                        cols_list = ['"Index"', '"PK_number"'] + [
                            f'"{c}"' for c in data.keys()
                        ]
                        placeholders = ', '.join('?' * (len(data) + 2))
                        values = [next_index, pk] + list(data.values())
                        cursor.execute(
                            f'INSERT INTO "{table}" ({", ".join(cols_list)}) '
                            f'VALUES ({placeholders})',
                            values,
                        )
                    next_index += 1
                    stats['added'] += 1

            except Exception as e:
                stats['error'] += 1
                if stats['error'] <= 3:
                    self.stderr.write(self.style.WARNING(
                        f'    [{sheet_name}] 행 {count} 오류: {e}'))

        return stats
