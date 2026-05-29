import json
import re
import threading
from datetime import date, datetime

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import login, get_user_model
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db import transaction
from django.db.models import Count, Q, Avg
from django.http import JsonResponse, HttpResponseBadRequest
from django.utils import timezone
from django.views.decorators.http import require_POST, require_GET

from .models import (
    WritingUnit, WritingProblem, UnitAssignment,
    WritingSession, WritingAttempt,
    StudentProfile, Achievement, StudentAchievement,
    StudentUnitLevel,
    BugReport,
    DailyStudyGoal,
)
from .services.excel import parse_writing_excel, parse_filename
from .services.students_excel import parse_students_excel
from .services.ai import generate_word_hints, generate_word_hints_batch
from .services import scoring, level as level_service


# 진행 상태 추적 (메모리, 프로세스당)
_hint_progress = {}  # {unit_id: {'total', 'completed', 'done', 'running'}}
_hint_progress_lock = threading.Lock()


# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────

def is_teacher(user):
    if not user.is_authenticated:
        return False
    if user.is_superuser or user.is_staff:
        return True
    if hasattr(user, 'member_type'):
        return user.member_type in ('academy_admin', 'admin')
    if hasattr(user, 'is_academy') and user.is_academy:
        return True
    return False


def teacher_required(view_func):
    return login_required(user_passes_test(is_teacher, login_url='/login/')(view_func))


def get_or_create_profile(user):
    profile, _ = StudentProfile.objects.get_or_create(student=user)
    return profile


def update_login_streak(profile):
    today = date.today()
    last = profile.last_login_date
    if last is None:
        profile.login_streak_days = 1
    elif last == today:
        return  # 이미 오늘 처리됨
    elif (today - last).days == 1:
        profile.login_streak_days += 1
    else:
        profile.login_streak_days = 1
    profile.last_login_date = today
    profile.save(update_fields=['login_streak_days', 'last_login_date'])


# ─────────────────────────────────────────────
# 데모 (시연용 자동 로그인)
# ─────────────────────────────────────────────

def demo_login(request):
    """샘플 시연용 — 로그인 없이 데모 계정으로 자동 진입 + 활성 단원 자동 배정."""
    Member = get_user_model()
    user, _ = Member.objects.get_or_create(
        login_id='demo',
        defaults={
            'username': '데모',
            'member_type': 'user',
            'is_active': True,
            'is_approved': True,
        },
    )
    # 기존 계정이 비활성/미승인일 수 있으니 강제로 맞춰둠
    if not (user.is_active and user.is_approved):
        user.is_active = True
        user.is_approved = True
        user.save(update_fields=['is_active', 'is_approved'])

    # 활성 단원 전부 자동 배정 (중복은 무시)
    active_units = WritingUnit.objects.filter(is_active=True)
    UnitAssignment.objects.bulk_create(
        [UnitAssignment(student=user, unit=u) for u in active_units],
        ignore_conflicts=True,
    )

    user.backend = 'member.backends.LoginIdOrEmailBackend'
    login(request, user)
    return redirect('writing:home')


# ─────────────────────────────────────────────
# 학생 화면들
# ─────────────────────────────────────────────

@login_required
def student_home(request):
    """영작훈련 학생 홈 — 배정된 단원 목록 (선생님은 전체)"""
    profile = get_or_create_profile(request.user)
    update_login_streak(profile)

    # 일반 학생인데 학원이 재원생으로 승인 안 했으면 안내 페이지
    if not is_teacher(request.user) and not getattr(request.user, 'is_approved', False):
        return render(request, 'writing/student_pending.html', {})

    if is_teacher(request.user):
        # 선생님/관리자는 전체 단원 풀이 가능
        units = WritingUnit.objects.filter(is_active=True).order_by('-created_at')
        is_assigned_view = False
    else:
        # 일반 학생: 배정된 단원만
        assignments = UnitAssignment.objects.filter(student=request.user).select_related('unit')
        units = [a.unit for a in assignments if a.unit.is_active]
        is_assigned_view = True

    # 학생의 단원별 숙련 Lv 일괄 조회 (학생 화면만 적용 — 선생님은 정렬에 영향 X)
    if is_assigned_view:
        level_map = {
            sul.unit_id: sul.level
            for sul in StudentUnitLevel.objects.filter(student=request.user)
        }
    else:
        level_map = {}

    # 각 단원마다 학생의 진행 상태
    unit_info = []
    for unit in units:
        sessions = WritingSession.objects.filter(student=request.user, unit=unit)
        last = sessions.order_by('-started_at').first()
        u_level = level_map.get(unit.id, 1)
        unit_info.append({
            'unit': unit,
            'attempts_count': sessions.count(),
            'last_score': last.total_score if last else None,
            'last_started': last.started_at if last else None,
            'unit_level': u_level,
            'level_info': level_service.level_summary(u_level),
        })

    # 학생 화면만 Lv 오름차순 정렬 (약한 단원 위로 — 우선 학습 유도).
    # 동일 Lv 안에서는 안 풀어본 단원 우선, 그 다음 오래 못 풀어본 단원.
    if is_assigned_view:
        unit_info.sort(key=lambda x: (
            x['unit_level'],
            0 if x['last_started'] is None else 1,
            x['last_started'].timestamp() if x['last_started'] else 0,
        ))

    return render(request, 'writing/home.html', {
        'profile': profile,
        'unit_info': unit_info,
        'is_assigned_view': is_assigned_view,
    })


@login_required
def flashcard_view(request, unit_id):
    """플래쉬카드 학습 화면 — 한글 ↔ 영어 뒤집기."""
    unit = get_object_or_404(WritingUnit, pk=unit_id, is_active=True)
    if not is_teacher(request.user):
        if not UnitAssignment.objects.filter(student=request.user, unit=unit).exists():
            messages.error(request, '이 단원은 배정되지 않았습니다.')
            return redirect('writing:home')
    problems = list(unit.problems.all().order_by('index'))
    cards = [{'index': p.index, 'korean': p.korean, 'english': p.english} for p in problems]
    return render(request, 'writing/flashcard.html', {
        'unit': unit,
        'cards_json': json.dumps(cards, ensure_ascii=False),
        'total': len(cards),
    })


@login_required
def start_session(request, unit_id):
    """단원 풀이 시작 — WritingSession 생성 후 풀이 화면으로.

    ?view=1 이면 '보고 학습 모드' (점수 ×0.5, 단계 재계산 X).
    """
    unit = get_object_or_404(WritingUnit, pk=unit_id, is_active=True)

    # 권한 체크: 일반 학생이면 배정받았는지
    if not is_teacher(request.user):
        if not UnitAssignment.objects.filter(student=request.user, unit=unit).exists():
            messages.error(request, '이 단원은 배정되지 않았습니다. 선생님께 문의하세요.')
            return redirect('writing:home')

    if not unit.problems.exists():
        messages.error(request, '이 단원에 문제가 없습니다.')
        return redirect('writing:home')

    view_mode = request.GET.get('view') == '1'
    session = WritingSession.objects.create(
        student=request.user, unit=unit, view_mode=view_mode,
    )
    return redirect('writing:session', session_id=session.id)


IGNORED_WORDS = {'a', 'an', 'the'}

# 약자 패턴 (Mt., Dr., Mr., Mrs., Ms., St., Jr., Sr., Prof., Inc., Ltd., Co., vs., etc.)
_ABBREV_PATTERN = re.compile(
    r'^(Mt|Dr|Mr|Mrs|Ms|St|Jr|Sr|Prof|Inc|Ltd|Co|vs|etc|Ave|Blvd|Rd)\.?$',
    re.IGNORECASE,
)
# 숫자 (1815, 5, 1st, 2nd, 1990s, 7,000 / 1,500,000 등 — 콤마 포함)
_NUMBER_PATTERN = re.compile(r'^\d+(st|nd|rd|th|s)?$', re.IGNORECASE)
_NUMBER_WITH_COMMA_PATTERN = re.compile(r'^\d{1,3}(,\d{3})+(st|nd|rd|th|s)?$', re.IGNORECASE)

_PUNCT = ',.!?;:"\'()[]{}'


def _should_auto_fill(word, position):
    """
    학습 가치가 낮아 자동으로 채워줄 단어인지 판정.
    position: 문장 내 단어 인덱스 (0부터)
    """
    if not word:
        return False
    cleaned = word.strip(_PUNCT)
    if not cleaned:
        return False

    # 관사
    if cleaned.lower() in IGNORED_WORDS:
        return True
    # 숫자 (서수/연도 + 콤마 포함 숫자 7,000 / 1,500,000 등)
    if _NUMBER_PATTERN.match(cleaned):
        return True
    if _NUMBER_WITH_COMMA_PATTERN.match(cleaned):
        return True
    # 약자 (Mt., Dr. 등)
    if _ABBREV_PATTERN.match(cleaned):
        return True
    # 모두 대문자 (USA, AI, NASA 등) — 2글자 이상은 약자/축약으로 간주
    if len(cleaned) >= 2 and cleaned.isupper():
        return True
    # (이전엔 "중간 대문자 시작 = 고유명사" 룰이 있었지만, 제목 케이스 데이터에선
    # 일반 명사도 대문자라 잘못 자동 채우기 됐음 — 제거)
    return False


# 하위 호환 — 다른 곳에서 쓰일 수 있으므로 유지
def _is_ignored_word(word):
    cleaned = word.strip(_PUNCT).lower()
    return cleaned in IGNORED_WORDS


@login_required
def session_view(request, session_id):
    """실제 풀이 화면"""
    session = get_object_or_404(WritingSession, pk=session_id)

    # 본인 세션만
    if session.student != request.user:
        messages.error(request, '본인 세션이 아닙니다.')
        return redirect('writing:home')

    # 이미 완료된 세션은 결과 화면으로
    if session.finished_at:
        return redirect('writing:result', session_id=session.id)

    profile = get_or_create_profile(request.user)
    problems = list(session.unit.problems.all().order_by('index'))

    # 이 세션 동안 적용될 단원 Lv (세션 시작 시 frozen — 풀이 중 흔들리지 않게)
    is_preview = is_teacher(session.student)
    if is_preview:
        unit_level = 1
    else:
        unit_level = level_service.get_or_create_unit_level(session.student, session.unit).level
    unit_level_info = level_service.level_summary(unit_level)

    # 클라이언트에 보낼 문제 데이터
    # 관사/고유명사/숫자/약자는 자동 채우기 — 정답값을 노출하지만 학습 가치 낮은 단어들
    problems_data = []
    for p in problems:
        words = p.english_words
        hints = p.word_hints or []
        word_meta = []
        for i, w in enumerate(words):
            hint = hints[i] if i < len(hints) and isinstance(hints[i], dict) else {}
            is_proper = bool(hint.get('proper_noun'))
            if is_proper or _should_auto_fill(w, i):
                word_meta.append({'auto': True, 'value': w})
            else:
                word_meta.append({'auto': False})
        problems_data.append({
            'id': p.id,
            'index': p.index,
            'korean': p.korean,
            'word_count': len(words),
            'words': word_meta,
        })

    return render(request, 'writing/session.html', {
        'session': session,
        'unit': session.unit,
        'profile': profile,
        'problems_json': json.dumps(problems_data, ensure_ascii=False),
        'total_problems': len(problems),
        'unit_level': unit_level,
        'unit_level_info': unit_level_info,
        'unit_level_json': json.dumps(unit_level_info, ensure_ascii=False),
        'view_mode': session.view_mode,
    })


@login_required
def result_view(request, session_id):
    """결과 화면"""
    session = get_object_or_404(WritingSession, pk=session_id)
    if session.student != request.user:
        return redirect('writing:home')

    profile = get_or_create_profile(request.user)
    attempts = WritingAttempt.objects.filter(session=session)
    total_words_attempted = attempts.values('problem_id', 'word_index').distinct().count()
    correct_first_try = attempts.filter(is_correct=True, attempt_num=1).count()
    accuracy = (correct_first_try / total_words_attempted * 100) if total_words_attempted else 0

    earned_badges = StudentAchievement.objects.filter(
        student=request.user,
        earned_at__gte=session.started_at,
    ).select_related('achievement')

    elapsed_seconds = 0
    if session.finished_at:
        elapsed_seconds = int((session.finished_at - session.started_at).total_seconds())
    target_seconds = session.unit.computed_target_seconds
    within_target = session.finished_at and elapsed_seconds <= target_seconds

    return render(request, 'writing/result.html', {
        'session': session,
        'unit': session.unit,
        'profile': profile,
        'accuracy': round(accuracy, 1),
        'total_words': total_words_attempted,
        'correct_first_try': correct_first_try,
        'earned_badges': earned_badges,
        'elapsed_seconds': elapsed_seconds,
        'target_seconds': target_seconds,
        'within_target': within_target,
        'time_bonus': session.time_bonus_earned,
    })


# ─────────────────────────────────────────────
# AJAX API — 단어 채점
# ─────────────────────────────────────────────

@login_required
@require_POST
def check_word_api(request):
    """
    단어 채점 + 점수/콤보/XP 처리.

    POST body (JSON):
      session_id, problem_id, word_index, input, time_taken_seconds

    Response:
      is_correct, attempt_num, hint_level, word_done, next_hint,
      score_earned, combo_bonus,
      current_word_combo, total_score, total_xp, level, level_up,
      badges_earned
    """
    try:
        data = json.loads(request.body)
        session_id = int(data['session_id'])
        problem_id = int(data['problem_id'])
        word_index = int(data['word_index'])
        student_input = str(data.get('input', ''))
        time_taken = int(data.get('time_taken_seconds', 0))
        is_auto_fill = bool(data.get('auto', False))
    except (json.JSONDecodeError, KeyError, ValueError):
        return HttpResponseBadRequest('Invalid request')

    session = get_object_or_404(WritingSession, pk=session_id)
    if session.student != request.user:
        return JsonResponse({'error': 'forbidden'}, status=403)
    if session.finished_at:
        return JsonResponse({'error': 'session_finished'}, status=400)

    # 관리자/선생님의 "미리 풀어보기" — 학생 프로필·XP·배지·리더보드에 영향 X
    is_preview = is_teacher(session.student)

    problem = get_object_or_404(WritingProblem, pk=problem_id, unit=session.unit)
    words = problem.english_words
    if word_index < 0 or word_index >= len(words):
        return HttpResponseBadRequest('Invalid word_index')

    correct_word = words[word_index]

    # 자동 채우기는 서버 측에서도 그 단어가 실제 자동 대상인지 검증 (악용 방지)
    if is_auto_fill:
        hints = problem.word_hints or []
        hint = hints[word_index] if word_index < len(hints) and isinstance(hints[word_index], dict) else {}
        is_proper = bool(hint.get('proper_noun'))
        if not is_proper and not _should_auto_fill(correct_word, word_index):
            is_auto_fill = False

    # 이 단어에 대한 이전 시도 횟수
    prev_attempts = WritingAttempt.objects.filter(
        session=session, problem=problem, word_index=word_index
    )
    prev_count = prev_attempts.count()

    # 이미 done(맞췄거나 3회 다 썼거나)이면 무시
    last_done = prev_attempts.filter(Q(is_correct=True) | Q(hint_level=3)).exists()
    if last_done:
        return JsonResponse({'error': 'word_already_done'}, status=400)

    attempt_num = prev_count + 1
    is_correct = scoring.check_correctness(student_input, correct_word)

    profile = get_or_create_profile(request.user)

    # 단원 Lv — 세션 동안 freeze (recompute는 세션 완료 시). 미리보기는 항상 Lv1.
    if is_preview:
        unit_level = 1
    else:
        unit_level = level_service.get_or_create_unit_level(request.user, session.unit).level
    max_attempts = level_service.max_attempts_for(unit_level)
    xp_mult = level_service.xp_multiplier(unit_level)
    # 보고 학습 모드 — 그 세션의 모든 점수에 0.5 곱하기
    view_mult = scoring.VIEW_MODE_MULTIPLIER if session.view_mode else 1.0

    score_earned = 0
    combo_bonus = 0
    hint_level_to_show = 0
    next_hint = None
    word_done = False
    fully_failed = False

    if is_correct:
        # 정답
        word_done = True
        hint_level_to_show = attempt_num - 1  # 0/1/2
        if is_auto_fill:
            # 자동 채우기는 점수/콤보 없음 (학생이 노력한 게 아님)
            score_earned = 0
        else:
            base_score = scoring.calculate_word_score(attempt_num, time_taken)
            score_earned = round(base_score * xp_mult * view_mult)
            if is_preview:
                # 관리자 미리보기 — 프로필/콤보에 영향 X (점수는 세션 표시용으로 유지)
                pass
            else:
                new_combo, hit_milestone = scoring.update_word_combo(profile, True, False)
                if hit_milestone:
                    combo_bonus = round(scoring.WORD_COMBO_BONUSES.get(new_combo, 0) * view_mult)
                score_earned += combo_bonus
    else:
        # 오답 — 그 Lv의 최대 시도 도달 시 정답 공개
        if attempt_num >= max_attempts:
            word_done = True
            fully_failed = True
            hint_level_to_show = 3
            score_earned = round(scoring.SCORE_REVEAL * xp_mult * view_mult)
            next_hint = scoring.get_hint_content(problem, word_index, 3)
            # 콤보 끊김
            if not is_preview:
                scoring.update_word_combo(profile, False, True)
        else:
            # 다음 힌트 보여줌 (Lv1: 1차→한글, 2차→첫글자 / Lv2: 1차→한글)
            hint_level_to_show = attempt_num
            score_earned = 0
            # 학생 입력이 정답의 70% 이상 prefix 매칭이면 정답 markup으로 (한글뜻 건너뜀)
            match_ratio = scoring.prefix_match_ratio(student_input, correct_word)
            if match_ratio >= scoring.NEAR_MISS_THRESHOLD:
                next_hint = {
                    'type': 'near_miss',
                    'content': correct_word,
                    'student': student_input.strip(),
                }
            else:
                next_hint = scoring.get_hint_content(problem, word_index, hint_level_to_show)

    # WritingAttempt 저장
    WritingAttempt.objects.create(
        session=session,
        problem=problem,
        word_index=word_index,
        input_value=student_input[:200],
        correct_answer=correct_word,
        hint_level=hint_level_to_show,
        is_correct=is_correct,
        attempt_num=attempt_num,
        time_taken_seconds=time_taken,
        score_earned=score_earned,
    )

    # 세션 점수 / 프로필 XP 업데이트 (단어마다, 가볍게)
    session.total_score += score_earned
    if profile.current_word_combo > session.max_word_combo:
        session.max_word_combo = profile.current_word_combo
    session.save(update_fields=['total_score', 'max_word_combo'])

    old_level = scoring.compute_level(profile.total_xp)
    new_level = old_level
    level_up = False
    if not is_preview:
        profile.total_xp += score_earned
        new_level = scoring.compute_level(profile.total_xp)
        level_up = new_level > old_level
        # update_fields로 필수 컬럼만 (배지/통계는 complete_problem에서 일괄)
        profile.save(update_fields=[
            'total_xp', 'current_word_combo', 'max_word_combo_ever',
        ])

    return JsonResponse({
        'is_correct': is_correct,
        'attempt_num': attempt_num,
        'hint_level': hint_level_to_show,
        'word_done': word_done,
        'next_hint': next_hint,
        'score_earned': score_earned,
        'combo_bonus': combo_bonus,
        'current_word_combo': profile.current_word_combo,
        'total_score': session.total_score,
        'total_xp': profile.total_xp,
        'level': new_level,
        'title': scoring.compute_title(new_level),
        'level_up': level_up,
        'xp_in_level': profile.xp_in_current_level,
        'badges_earned': [],  # 배지는 문장 끝에 일괄 처리
    })


@login_required
@require_POST
def complete_problem_api(request):
    """
    문제 1개 풀이 완료 시 (모든 단어 done) — 문장 콤보 처리.

    POST body: {session_id, problem_id}
    """
    try:
        data = json.loads(request.body)
        session_id = int(data['session_id'])
        problem_id = int(data['problem_id'])
    except (json.JSONDecodeError, KeyError, ValueError):
        return HttpResponseBadRequest('Invalid')

    session = get_object_or_404(WritingSession, pk=session_id)
    if session.student != request.user:
        return JsonResponse({'error': 'forbidden'}, status=403)
    is_preview = is_teacher(session.student)
    problem = get_object_or_404(WritingProblem, pk=problem_id, unit=session.unit)

    # 이 문제 단어 전체 시도 분석
    attempts = WritingAttempt.objects.filter(session=session, problem=problem)
    total_words = len(problem.english_words)
    distinct_words_done = attempts.values('word_index').distinct().count()
    if distinct_words_done < total_words:
        return JsonResponse({'error': 'not_all_words_done'}, status=400)

    # 사람이 실제로 맞춘 단어 수 (자동 채우기 = score_earned=0 제외)
    human_correct_count = attempts.filter(is_correct=True, score_earned__gt=0).count()

    # 무지성 통과 판정: 정답 0개 + 시도당 평균 시간이 너무 짧음
    # (자동 채우기는 is_correct=True AND score_earned=0 으로 식별)
    non_auto_attempts = attempts.exclude(is_correct=True, score_earned=0)
    non_auto_count = non_auto_attempts.count()
    avg_time = (
        non_auto_attempts.aggregate(Avg('time_taken_seconds'))['time_taken_seconds__avg'] or 0
    )
    forfeit = (
        human_correct_count == 0
        and non_auto_count > 0
        and avg_time < scoring.NO_THINK_THRESHOLD_SEC
    )

    # 무실수 문장 = 모든 단어가 1차 시도 정답 (단, 무지성 케이스는 제외)
    all_first_try = (
        not forfeit
        and attempts.filter(is_correct=True, attempt_num=1).count() == total_words
    )

    profile = get_or_create_profile(request.user)

    forfeit_amount = 0
    if forfeit:
        # 이 문장에서 얻은 모든 점수 회수 (자동 채우기 +0은 영향 없음)
        forfeit_amount = sum(a.score_earned for a in attempts)
        if forfeit_amount > 0:
            session.total_score = max(0, session.total_score - forfeit_amount)
            if not is_preview:
                profile.total_xp = max(0, profile.total_xp - forfeit_amount)
        # 문장 콤보도 끊김
        if not is_preview:
            scoring.update_sentence_combo(profile, False)
        new_sent_combo, milestone = 0, False
    else:
        if is_preview:
            new_sent_combo, milestone = 0, False
        else:
            new_sent_combo, milestone = scoring.update_sentence_combo(profile, all_first_try)

    extra_score = 0
    perfect_bonus = 0
    sentence_combo_bonus = 0
    view_mult = scoring.VIEW_MODE_MULTIPLIER if session.view_mode else 1.0
    if all_first_try:
        perfect_bonus = round(scoring.PERFECT_SENTENCE_BONUS * view_mult)
        session.perfect_sentences += 1
        if milestone:
            sentence_combo_bonus = round(scoring.SENTENCE_COMBO_BONUSES.get(new_sent_combo, 0) * view_mult)
        extra_score = perfect_bonus + sentence_combo_bonus
        session.total_score += extra_score
        if not is_preview:
            profile.total_xp += extra_score

    if profile.current_sentence_combo > session.max_sentence_combo:
        session.max_sentence_combo = profile.current_sentence_combo

    # ── 배지 체크 (문장 단위, 단어마다 안 함) ──
    newly_earned = []
    if not is_preview:
        perfect_count_total = WritingAttempt.objects.filter(
            session__student=request.user,
            is_correct=True,
            attempt_num=1,
        ).count()

        earned_codes = scoring.check_badges(profile, {
            'is_correct_first_try': all_first_try,
            'perfect_count_total': perfect_count_total,
            'was_perfect_sentence': all_first_try,
            'was_perfect_unit': False,
            'speed_bonus_count': 0,
            'current_hour': datetime.now().hour,
        })

        if earned_codes:
            for code in earned_codes:
                ach = Achievement.objects.filter(code=code).first()
                if ach:
                    sa, created = StudentAchievement.objects.get_or_create(
                        student=request.user, achievement=ach,
                    )
                    if created:
                        newly_earned.append({
                            'icon': ach.icon, 'name': ach.name, 'description': ach.description,
                        })

    session.save()
    if not is_preview:
        profile.save()

    # 이 문장의 점수 비율 계산 (자동 채우기 제외, base 점수만 — speed/콤보 보너스 제외)
    # 자동 채우기 판정: _should_auto_fill (관사/약자/숫자/대문자약자) + word_hints.proper_noun
    words = problem.english_words
    hints = problem.word_hints or []
    def _is_auto(i, w):
        if _should_auto_fill(w, i):
            return True
        if i < len(hints) and isinstance(hints[i], dict) and hints[i].get('proper_noun'):
            return True
        return False
    non_auto_count = sum(1 for i, w in enumerate(words) if not _is_auto(i, w))
    sentence_max = non_auto_count * scoring.SCORE_BY_HINT_LEVEL[0]  # 단어당 10점 만점

    sentence_earned = 0
    for a in attempts:
        if a.is_correct and a.score_earned == 0:
            continue  # 자동 채우기
        if a.is_correct:
            # PERFECT(10) / GREAT(6) / GOOD(3) — base만
            sentence_earned += scoring.SCORE_BY_HINT_LEVEL.get(a.hint_level, 0)
        elif a.hint_level == 3:
            # BOO~ (정답 공개) +1
            sentence_earned += scoring.SCORE_REVEAL

    if forfeit:
        sentence_earned = 0
    sentence_pct = (sentence_earned / sentence_max * 100) if sentence_max > 0 else 100
    passed = sentence_pct >= scoring.SENTENCE_PASS_THRESHOLD

    return JsonResponse({
        'was_perfect_sentence': all_first_try,
        'perfect_bonus': perfect_bonus,
        'sentence_combo_bonus': sentence_combo_bonus,
        'forfeit': forfeit,
        'forfeit_amount': forfeit_amount,
        'avg_time_per_attempt': round(avg_time, 1),
        'sentence_score': sentence_earned,
        'sentence_max': sentence_max,
        'sentence_pct': round(sentence_pct, 1),
        'passed': passed,
        'pass_threshold': scoring.SENTENCE_PASS_THRESHOLD,
        'current_sentence_combo': profile.current_sentence_combo,
        'badges_earned': newly_earned,
        'total_score': session.total_score,
        'total_xp': profile.total_xp,
        'level': scoring.compute_level(profile.total_xp),
        'title': scoring.compute_title(scoring.compute_level(profile.total_xp)),
        'xp_in_level': profile.xp_in_current_level,
    })


@login_required
@require_POST
def reset_problem_api(request):
    """문제 재시도 — 그 문제의 attempts 삭제 + 점수 회수"""
    try:
        data = json.loads(request.body)
        session_id = int(data['session_id'])
        problem_id = int(data['problem_id'])
    except (json.JSONDecodeError, KeyError, ValueError):
        return HttpResponseBadRequest('Invalid')

    session = get_object_or_404(WritingSession, pk=session_id)
    if session.student != request.user:
        return JsonResponse({'error': 'forbidden'}, status=403)
    if session.finished_at:
        return JsonResponse({'error': 'session_finished'}, status=400)
    is_preview = is_teacher(session.student)

    attempts = WritingAttempt.objects.filter(session=session, problem_id=problem_id)
    forfeit = sum(a.score_earned for a in attempts)
    attempts.delete()

    profile = get_or_create_profile(request.user)
    if forfeit > 0:
        session.total_score = max(0, session.total_score - forfeit)
        session.save(update_fields=['total_score'])
        if not is_preview:
            profile.total_xp = max(0, profile.total_xp - forfeit)
            profile.save(update_fields=['total_xp'])

    # 콤보도 끊김
    if not is_preview:
        profile.current_word_combo = 0
        profile.current_sentence_combo = 0
        profile.save(update_fields=['current_word_combo', 'current_sentence_combo'])

    return JsonResponse({
        'success': True,
        'forfeit': forfeit,
        'total_score': session.total_score,
        'total_xp': profile.total_xp,
        'xp_in_level': profile.xp_in_current_level,
        'level': scoring.compute_level(profile.total_xp),
        'title': scoring.compute_title(scoring.compute_level(profile.total_xp)),
    })


TIME_BONUS = 100  # baseline(단어당 7초) 깨면 주는 보너스


@login_required
@require_POST
def complete_session_api(request):
    """세션 완료 + 도전 baseline 안 완료 시 보너스 + 시간 리더보드 진입."""
    try:
        data = json.loads(request.body)
        session_id = int(data['session_id'])
    except (json.JSONDecodeError, KeyError, ValueError):
        return HttpResponseBadRequest('Invalid')

    session = get_object_or_404(WritingSession, pk=session_id)
    if session.student != request.user:
        return JsonResponse({'error': 'forbidden'}, status=403)
    is_preview = is_teacher(session.student)

    time_bonus = 0
    elapsed = 0
    baseline = session.unit.computed_target_seconds

    view_mult = scoring.VIEW_MODE_MULTIPLIER if session.view_mode else 1.0
    if not session.finished_at:
        session.finished_at = timezone.now()
        elapsed = int((session.finished_at - session.started_at).total_seconds())
        if elapsed <= baseline:
            time_bonus = round(TIME_BONUS * view_mult)
            session.total_score += time_bonus
            session.time_bonus_earned = time_bonus
            if not is_preview:
                profile = get_or_create_profile(request.user)
                profile.total_xp += time_bonus
                profile.save(update_fields=['total_xp'])
        session.save(update_fields=['finished_at', 'total_score', 'time_bonus_earned'])
    else:
        elapsed = int((session.finished_at - session.started_at).total_seconds())
        time_bonus = session.time_bonus_earned

    # 단원 단계 재계산 — 보고 학습 모드는 평가에서 제외
    old_level = new_level = 1
    if not is_preview and not session.view_mode:
        old_level, new_level = level_service.recompute_unit_level(
            request.user, session.unit, session
        )
    elif not is_preview:
        # view_mode면 현재 단계 그대로 응답 (UI 갱신용)
        sul = level_service.get_or_create_unit_level(request.user, session.unit)
        old_level = new_level = sul.level

    return JsonResponse({
        'success': True,
        'time_bonus': time_bonus,
        'elapsed_seconds': elapsed,
        'baseline_seconds': baseline,
        'broke_baseline': elapsed <= baseline,
        'unit_level_old': old_level,
        'unit_level_new': new_level,
        'unit_level_changed': old_level != new_level,
        'redirect_url': f'/training/writing/result/{session.id}/',
    })


@login_required
@require_GET
def leaderboard_api(request, unit_id):
    """단원별 리더보드 — (1) 점수 Top 3 (2) baseline 깬 학생의 시간 Top 3."""
    from django.contrib.auth import get_user_model
    from django.db.models import Max
    unit = get_object_or_404(WritingUnit, pk=unit_id)
    baseline = unit.computed_target_seconds

    # 관리자/선생님 세션은 리더보드에서 제외
    finished_sessions = WritingSession.objects.filter(
        unit=unit, finished_at__isnull=False,
        student__is_staff=False,
        student__is_superuser=False,
        student__is_academy=False,
    )

    # ── 점수 리더보드 ──
    best_score_per_student = (
        finished_sessions
        .values('student_id')
        .annotate(best=Max('total_score'))
        .order_by('-best', 'student_id')
    )
    score_top_qs = list(best_score_per_student[:3])

    User = get_user_model()
    score_ids = [r['student_id'] for r in score_top_qs]
    users_by_id = {u.id: u for u in User.objects.filter(pk__in=score_ids)}

    score_top = []
    for rank, r in enumerate(score_top_qs, start=1):
        u = users_by_id.get(r['student_id'])
        if not u:
            continue
        score_top.append({
            'rank': rank,
            'name': u.username or f'#{u.id}',
            'score': r['best'],
            'is_me': u.id == request.user.id,
        })

    me_best_score = (
        finished_sessions
        .filter(student=request.user)
        .aggregate(best=Max('total_score'))['best']
    )
    me_score_rank = None
    if me_best_score is not None:
        higher = best_score_per_student.filter(best__gt=me_best_score).count()
        me_score_rank = higher + 1

    # ── 시간 리더보드 (baseline 깬 학생만 등록) ──
    student_best_time = {}  # student_id -> {'name', 'duration'}
    for s in finished_sessions.select_related('student'):
        duration = int((s.finished_at - s.started_at).total_seconds())
        if duration > baseline:
            continue
        existing = student_best_time.get(s.student_id)
        if existing is None or duration < existing['duration']:
            student_best_time[s.student_id] = {
                'student_id': s.student_id,
                'name': s.student.username or f'#{s.student_id}',
                'duration': duration,
            }
    time_sorted = sorted(student_best_time.values(), key=lambda x: x['duration'])
    time_top = []
    for rank, r in enumerate(time_sorted[:3], start=1):
        time_top.append({
            'rank': rank,
            'name': r['name'],
            'seconds': r['duration'],
            'is_me': r['student_id'] == request.user.id,
        })

    me_time_entry = student_best_time.get(request.user.id)
    me_time_rank = None
    if me_time_entry:
        me_time_rank = next(
            (i + 1 for i, r in enumerate(time_sorted) if r['student_id'] == request.user.id),
            None,
        )

    return JsonResponse({
        'unit_id': unit.id,
        'unit_title': unit.title,
        'baseline_seconds': baseline,
        'total_words': unit.total_words,
        'score': {
            'top': score_top,
            'me_rank': me_score_rank,
            'me_best': me_best_score,
        },
        'time': {
            'top': time_top,
            'me_rank': me_time_rank,
            'me_best_seconds': me_time_entry['duration'] if me_time_entry else None,
        },
    })


# ─────────────────────────────────────────────
# 선생님 화면들 (기존)
# ─────────────────────────────────────────────

VALID_GRADES = {g[0] for g in WritingUnit.GRADE_CHOICES}


def _start_hint_generation(unit_id, force=False):
    """단원의 AI 한글뜻 생성 백그라운드 스레드 시작.
    force=True면 이미 한글뜻 있는 문제도 다시 생성 (고유명사 식별 보완용)."""
    unit = WritingUnit.objects.filter(pk=unit_id).first()
    if not unit:
        return False
    if force:
        WritingProblem.objects.filter(unit_id=unit_id).update(word_hints=[])
    remaining = unit.problems.filter(word_hints=[]).count()
    if remaining == 0:
        return False
    with _hint_progress_lock:
        existing = _hint_progress.get(unit_id)
        if existing and existing.get('running'):
            return False
        _hint_progress[unit_id] = {
            'total': remaining, 'completed': 0,
            'done': False, 'running': True, 'error': None,
        }
    threading.Thread(target=_run_hint_generation, args=(unit_id,), daemon=True).start()
    return True


@teacher_required
def upload_view(request):
    if request.method != 'POST':
        return render(request, 'writing/upload.html', {})

    files = request.FILES.getlist('excel_file')
    if not files:
        messages.error(request, '엑셀 파일을 1개 이상 선택해주세요.')
        return render(request, 'writing/upload.html', {})

    created_units = []
    total_problems = 0
    total_skipped = 0
    file_errors = []
    duplicates = []

    for f in files:
        if not f.name.lower().endswith(('.xlsx', '.xls')):
            file_errors.append(f'{f.name}: xlsx/xls만 업로드 가능')
            continue
        if f.size > 10 * 1024 * 1024:
            file_errors.append(f'{f.name}: 10MB 초과')
            continue

        meta = parse_filename(f.name)
        title = meta['title'] or re.sub(r'\.[^.]+$', '', f.name)
        grade = meta['grade'] if meta['grade'] in VALID_GRADES else '기타'
        publisher = meta['publisher']

        # 중복 체크: 학년 + 출판사 + 단원명이 모두 같으면 skip
        if WritingUnit.objects.filter(title=title, grade=grade, publisher=publisher).exists():
            duplicates.append(f'{f.name} → "{title}" ({grade} / {publisher or "출판사 없음"}) 이미 존재')
            continue

        result = parse_writing_excel(f)
        if not result['success']:
            file_errors.append(f'{f.name}: {"; ".join(result["errors"])}')
            continue

        try:
            with transaction.atomic():
                unit = WritingUnit.objects.create(
                    title=title,
                    publisher=publisher,
                    grade=grade,
                    created_by=request.user,
                )
                for p in result['problems']:
                    WritingProblem.objects.create(
                        unit=unit,
                        index=p['index'],
                        korean=p['korean'],
                        english=p['english'],
                        word_hints=[],
                    )
            created_units.append(unit)
            total_problems += len(result['problems'])
            total_skipped += result.get('skipped_short', 0)
        except Exception as e:
            file_errors.append(f'{f.name}: 저장 실패 — {e}')

    for err in file_errors:
        messages.warning(request, err)
    for dup in duplicates:
        messages.info(request, f'중복 skip: {dup}')

    if not created_units:
        if duplicates:
            messages.warning(request, '선택한 파일 모두 이미 등록된 단원이라 생성된 단원이 없습니다.')
        else:
            messages.error(request, '생성된 단원이 없습니다.')
        return render(request, 'writing/upload.html', {})

    # 새로 생긴 단원들에 대해 AI 한글뜻 생성 자동 시작
    hint_started = 0
    for u in created_units:
        if _start_hint_generation(u.id):
            hint_started += 1

    skip_msg = f' · 3단어 이하 제외 {total_skipped}행' if total_skipped else ''
    dup_msg = f' · 중복 skip {len(duplicates)}개' if duplicates else ''
    hint_msg = f' · AI 한글뜻 생성 {hint_started}개 단원 자동 시작 (단원 관리 표에서 진행도 확인)' if hint_started else ''
    messages.success(
        request,
        f'단원 {len(created_units)}개 생성 · 문제 {total_problems}개 등록{skip_msg}{dup_msg}{hint_msg}',
    )

    if len(created_units) == 1:
        return redirect('writing:unit_detail', unit_id=created_units[0].id)
    return redirect('writing:unit_list')


_GRADES_ORDER = [
    '초1', '초2', '초3', '초4', '초5', '초6',
    '중1', '중2', '중3', '고1', '고2', '고3',
]

# 단원 title 앞머리에서 '과' 키 추출 — 내신을 과별로 묶기 위함
_LESSON_KEY_RE = re.compile(
    r'^\s*((?:\d+\s*-\s*\d+|\d+)\s*과|\d+\s*단원|Lesson\s*\d+)',
    re.IGNORECASE,
)


def _lesson_key(title: str):
    """제목에서 '1과', '1-2과', '3단원', 'Lesson 5' 같은 키를 뽑는다. 없으면 None."""
    if not title:
        return None
    m = _LESSON_KEY_RE.match(title)
    if not m:
        return None
    return re.sub(r'\s+', ' ', m.group(1)).strip()


def _lesson_sort_key(label: str):
    """과 라벨을 자연 정렬용 튜플로 변환. '(과 미분류)'는 맨 뒤."""
    if not label or label == '(과 미분류)':
        return (1, 9999, 9999, '')
    nums = re.findall(r'\d+', label)
    if not nums:
        return (0, 9998, 9998, label)
    primary = int(nums[0])
    secondary = int(nums[1]) if len(nums) > 1 else 0
    return (0, primary, secondary, label)


# 학교 내신 시험기 패턴: "(YYYY년) 학교명+학년 N학기 (중간/기말/모의)고사"
#   예: "2026년 동백중3 1학기 기말고사" → period="2026년 1학기 기말고사", school="동백중3"
_SCHOOL_EXAM_PUB_RE = re.compile(
    r'^(?:(?P<year>\d{4})년\s+)?'
    r'(?P<school>[^\s_]+?(?:초[1-6]|중[1-3]|고[1-3]))\s+'
    r'(?P<term>\d학기)\s+'
    r'(?P<exam>중간고사|기말고사|중간|기말|모의고사|[가-힣A-Za-z]+고사)'
    r'(?:\s+(?P<suffix>.+))?$'
)


def _parse_school_exam_publisher(pub: str):
    """publisher 문자열에서 (시험기, 학교) 추출. 매치 안 되면 None."""
    if not pub:
        return None
    m = _SCHOOL_EXAM_PUB_RE.match(pub.strip())
    if not m:
        return None
    parts = []
    if m.group('year'):
        parts.append(f"{m.group('year')}년")
    parts.append(m.group('term'))
    parts.append(m.group('exam'))
    return ' '.join(parts), m.group('school')


def _unit_natural_key(unit):
    """단원명 안의 숫자로 자연 정렬 (외부지문 1, 외부지문 2, ..., 외부지문 10)."""
    nums = re.findall(r'\d+', unit.title or '')
    return (int(nums[-1]) if nums else 9999, unit.title or '')


def _category_key(title: str) -> str:
    """학교 내신 단원명에서 카테고리(prefix) 추출.

    예: "외부지문 1" → "외부지문", "본문 3" → "본문", "외부지문" → "외부지문"
    숫자 분리가 안 되면 단원명 그대로.
    """
    if not title:
        return '(분류없음)'
    t = title.strip()
    m = re.match(r'^(.+?)\s+\d+\s*$', t)
    if m:
        return m.group(1).strip()
    return t


@teacher_required
def unit_list(request):
    units = list(WritingUnit.objects.all().order_by('publisher', 'title'))
    with _hint_progress_lock:
        running_ids = {uid for uid, s in _hint_progress.items() if s.get('running')}
    for u in units:
        u.has_hints_count = u.problems.exclude(word_hints=[]).count()
        u.total_count = u.problems.count()
        u.hints_running = u.id in running_ids

    # 내신 그룹핑 — 두 줄기:
    #   (a) 시험기형: 학년 → 시험기(2026년 1학기 기말고사) → 학교(동백중3) → 단원
    #   (b) 일반 출판사형: 학년 → 출판사(NE능률 김성곤) → 과 → 단원
    # 부교재: 출판사 → 단원
    naesin_by_grade = {}   # grade → {'exam_periods': {period: {school: [units]}}, 'publishers': {pub: {lesson: [units]}}}
    buggyojae_by_pub = {}  # publisher → [units]
    for u in units:
        if u.grade in _GRADES_ORDER:
            bucket = naesin_by_grade.setdefault(
                u.grade, {'exam_periods': {}, 'publishers': {}}
            )
            parsed = _parse_school_exam_publisher(u.publisher)
            if parsed:
                period, school = parsed
                bucket['exam_periods'].setdefault(period, {}) \
                    .setdefault(school, {}) \
                    .setdefault(_category_key(u.title), []).append(u)
            else:
                lk = _lesson_key(u.title) or '(과 미분류)'
                bucket['publishers'].setdefault(u.publisher or '(미지정)', {}) \
                    .setdefault(lk, []).append(u)
        else:
            buggyojae_by_pub.setdefault(u.publisher or '(미지정)', []).append(u)

    naesin_groups = []
    for g in _GRADES_ORDER:
        if g not in naesin_by_grade:
            continue
        data = naesin_by_grade[g]
        pubs = []

        # 시험기 폴더 (최신 연도부터)
        for period in sorted(data['exam_periods'].keys(), reverse=True):
            schools_map = data['exam_periods'][period]
            schools = []
            for school_name in sorted(schools_map.keys()):
                cat_map = schools_map[school_name]
                categories = []
                for cat_name in sorted(cat_map.keys()):
                    us = sorted(cat_map[cat_name], key=_unit_natural_key)
                    categories.append({
                        'name': cat_name,
                        'units': us,
                        'count': len(us),
                    })
                schools.append({
                    'name': school_name,
                    'categories': categories,
                    'count': sum(c['count'] for c in categories),
                })
            pubs.append({
                'name': period,
                'is_exam_period': True,
                'schools': schools,
                'count': sum(s['count'] for s in schools),
            })

        # 일반 출판사 폴더 (이름 오름차순)
        for pub_name in sorted(data['publishers'].keys()):
            lesson_map = data['publishers'][pub_name]
            lessons = [
                {'name': lk, 'units': lesson_map[lk]}
                for lk in sorted(lesson_map.keys(), key=_lesson_sort_key)
            ]
            pubs.append({
                'name': pub_name,
                'is_exam_period': False,
                'lessons': lessons,
                'count': sum(len(l['units']) for l in lessons),
            })

        naesin_groups.append({
            'grade': g,
            'publishers': pubs,
            'count': sum(p['count'] for p in pubs),
        })

    buggyojae_groups = [
        {'name': name, 'units': us}
        for name, us in sorted(buggyojae_by_pub.items())
    ]

    return render(request, 'writing/unit_list.html', {
        'units': units,
        'naesin_groups': naesin_groups,
        'buggyojae_groups': buggyojae_groups,
        'naesin_total': sum(g['count'] for g in naesin_groups),
        'buggyojae_total': sum(len(g['units']) for g in buggyojae_groups),
    })


@teacher_required
@require_POST
def unit_delete(request):
    """체크박스로 고른 단원들을 cascade 데이터(문제·배정·세션·시도)와 함께 삭제."""
    raw_ids = request.POST.getlist('unit_ids')
    try:
        ids = [int(x) for x in raw_ids if x]
    except ValueError:
        ids = []

    if not ids:
        messages.warning(request, '삭제할 단원을 선택해주세요.')
        return redirect('writing:unit_list')

    qs = WritingUnit.objects.filter(pk__in=ids)
    count = qs.count()
    qs.delete()
    messages.success(request, f'단원 {count}개 삭제 완료.')
    return redirect('writing:unit_list')


@teacher_required
@require_POST
def problem_reorder(request, unit_id):
    """단원 문제 순서 일괄 갱신. body: {order: [pid1, pid2, ...]}."""
    unit = get_object_or_404(WritingUnit, pk=unit_id)
    try:
        data = json.loads(request.body or '{}')
        order = [int(x) for x in data.get('order', [])]
    except (json.JSONDecodeError, ValueError, TypeError):
        return HttpResponseBadRequest('Invalid')
    if not order:
        return HttpResponseBadRequest('Empty order')

    pids_in_unit = set(
        WritingProblem.objects.filter(unit=unit).values_list('id', flat=True)
    )
    filtered = [pid for pid in order if pid in pids_in_unit]

    with transaction.atomic():
        # unique_together 회피용 — 음수로 옮긴 뒤 재부여
        for offset, pid in enumerate(filtered):
            WritingProblem.objects.filter(pk=pid).update(index=-(offset + 1))
        for offset, pid in enumerate(filtered):
            WritingProblem.objects.filter(pk=pid).update(index=offset + 1)

    return JsonResponse({'success': True, 'count': len(filtered)})


@teacher_required
@require_POST
def problem_insert(request, unit_id):
    """단원에 새 문제 행 삽입. body: {after_index: int}.
    after_index=0 → 맨 위, after_index=N → N번째 뒤. 인덱스 자동 재정렬."""
    unit = get_object_or_404(WritingUnit, pk=unit_id)
    try:
        data = json.loads(request.body or '{}')
        after_index = int(data.get('after_index', 0))
    except (json.JSONDecodeError, ValueError, TypeError):
        return HttpResponseBadRequest('Invalid')

    with transaction.atomic():
        existing = list(
            WritingProblem.objects.filter(unit=unit).order_by('index', 'id')
        )
        new_pos = max(0, min(after_index, len(existing)))  # 0..N

        # unique_together(unit, index) 충돌 회피용 — 음수로 임시 옮기기
        for i, p in enumerate(existing):
            WritingProblem.objects.filter(pk=p.pk).update(index=-(i + 1))

        # 새 문제 위치 = new_pos + 1 (1-base) — 빈 셀로 시작
        new_problem = WritingProblem.objects.create(
            unit=unit,
            index=new_pos + 1,
            korean='',
            english='',
            word_hints=[],
        )

        # 나머지 재부여
        for offset, p in enumerate(existing):
            new_idx = offset + 1 if offset < new_pos else offset + 2
            WritingProblem.objects.filter(pk=p.pk).update(index=new_idx)

    return JsonResponse({'success': True, 'problem_id': new_problem.id, 'index': new_problem.index})


@teacher_required
@require_POST
def problems_delete(request):
    """문제 일괄 삭제. body: {problem_ids: [...]}. 삭제 후 같은 단원의 index를 1부터 재정렬."""
    try:
        data = json.loads(request.body)
        ids = [int(x) for x in data.get('problem_ids', [])]
    except (json.JSONDecodeError, ValueError, TypeError):
        return HttpResponseBadRequest('Invalid')
    if not ids:
        return HttpResponseBadRequest('Empty problem_ids')

    affected_unit_ids = list(
        WritingProblem.objects.filter(pk__in=ids).values_list('unit_id', flat=True).distinct()
    )
    with transaction.atomic():
        deleted, _ = WritingProblem.objects.filter(pk__in=ids).delete()
        # 단원별 index 재정렬 (1부터)
        for uid in affected_unit_ids:
            remaining = list(WritingProblem.objects.filter(unit_id=uid).order_by('index', 'id'))
            # unique_together(unit, index) 충돌 회피용 — 일단 음수로 옮긴 뒤 재부여
            for offset, p in enumerate(remaining, start=1):
                WritingProblem.objects.filter(pk=p.pk).update(index=-offset)
            for offset, p in enumerate(remaining, start=1):
                WritingProblem.objects.filter(pk=p.pk).update(index=offset)

    return JsonResponse({'success': True, 'deleted': deleted})


@teacher_required
@require_GET
def problem_hints_get(request, problem_id):
    """문제의 word_hints 조회 — 편집 모달용."""
    problem = get_object_or_404(WritingProblem, pk=problem_id)
    return JsonResponse({
        'id': problem.id,
        'english': problem.english,
        'korean': problem.korean,
        'word_hints': problem.word_hints or [],
    }, json_dumps_params={'ensure_ascii': False})


@teacher_required
@require_POST
def unit_replace_excel(request, unit_id):
    """기존 단원의 문제를 업로드한 엑셀로 통째 교체.

    POST multipart: excel_file
    동작: 기존 problems 전부 삭제 → 엑셀 파싱 결과로 재생성.
    AI 한글뜻은 비워짐 (필요 시 별도 일괄 생성).
    """
    unit = get_object_or_404(WritingUnit, pk=unit_id)
    f = request.FILES.get('excel_file')
    if not f:
        messages.error(request, '엑셀 파일을 선택해주세요.')
        return redirect('writing:unit_detail', unit_id=unit_id)
    if not f.name.lower().endswith(('.xlsx', '.xls')):
        messages.error(request, 'xlsx 또는 xls 파일만 가능합니다.')
        return redirect('writing:unit_detail', unit_id=unit_id)

    result = parse_writing_excel(f)
    if not result['success']:
        for err in result['errors']:
            messages.error(request, err)
        return redirect('writing:unit_detail', unit_id=unit_id)

    problems = result['problems']
    with transaction.atomic():
        WritingProblem.objects.filter(unit=unit).delete()
        WritingProblem.objects.bulk_create([
            WritingProblem(
                unit=unit,
                index=p['index'],
                korean=p['korean'],
                english=p['english'],
                word_hints=[],
            ) for p in problems
        ])

    parts = [f'{len(problems)}개 문제로 교체 완료']
    if result.get('skipped_short'):
        parts.append(f'3단어 이하 {result["skipped_short"]}개 자동 제외')
    parts.append('AI 한글뜻은 초기화됨 — 필요 시 재생성하세요')
    messages.success(request, ' · '.join(parts))
    return redirect('writing:unit_detail', unit_id=unit_id)


@teacher_required
@require_POST
def problem_update(request, problem_id):
    """문제 인라인 편집 — body: {korean?, english?, word_hints?} 중 보낸 필드만 갱신."""
    problem = get_object_or_404(WritingProblem, pk=problem_id)
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return HttpResponseBadRequest('Invalid JSON')

    update_fields = []
    if 'korean' in data:
        problem.korean = str(data['korean']).strip()
        update_fields.append('korean')
    if 'english' in data:
        problem.english = str(data['english']).strip()
        update_fields.append('english')
    if 'word_hints' in data:
        hints = data['word_hints']
        if not isinstance(hints, list):
            return HttpResponseBadRequest('word_hints must be a list')
        clean = []
        for h in hints:
            if not isinstance(h, dict):
                continue
            word = str(h.get('word', '')).strip()
            if not word:
                continue
            clean.append({
                'word': word,
                'meaning': str(h.get('meaning', '')).strip(),
                'proper_noun': bool(h.get('proper_noun', False)),
            })
        problem.word_hints = clean
        update_fields.append('word_hints')

    if not update_fields:
        return HttpResponseBadRequest('No fields to update')

    problem.save(update_fields=update_fields)
    return JsonResponse({
        'success': True,
        'id': problem.id,
        'korean': problem.korean,
        'english': problem.english,
        'word_hints': problem.word_hints,
    }, json_dumps_params={'ensure_ascii': False})


@teacher_required
def unit_detail(request, unit_id):
    unit = get_object_or_404(WritingUnit, pk=unit_id)
    problems = unit.problems.all().order_by('index')
    has_hints_count = sum(1 for p in problems if p.word_hints)
    total_count = problems.count()
    return render(request, 'writing/unit_detail.html', {
        'unit': unit,
        'problems': problems,
        'has_hints_count': has_hints_count,
        'total_count': total_count,
        'missing_hints_count': total_count - has_hints_count,
    })


BATCH_SIZE = 20  # API 호출당 문제 수


def _run_hint_generation(unit_id):
    """백그라운드 스레드에서 단원 전체 한글뜻 생성"""
    from django.db import connection
    try:
        problems = list(
            WritingProblem.objects.filter(unit_id=unit_id, word_hints=[]).order_by('index')
        )
        with _hint_progress_lock:
            _hint_progress[unit_id]['total'] = len(problems)
            _hint_progress[unit_id]['completed'] = 0

        for batch_start in range(0, len(problems), BATCH_SIZE):
            batch = problems[batch_start:batch_start + BATCH_SIZE]
            inputs = [{'english': p.english, 'korean': p.korean} for p in batch]
            hints_lists = generate_word_hints_batch(inputs)

            for problem, hints in zip(batch, hints_lists):
                if hints:
                    problem.word_hints = hints
                    problem.save(update_fields=['word_hints'])

            with _hint_progress_lock:
                _hint_progress[unit_id]['completed'] = batch_start + len(batch)

        with _hint_progress_lock:
            _hint_progress[unit_id]['done'] = True
            _hint_progress[unit_id]['running'] = False
    except Exception as e:
        print(f'[hint generation] unit {unit_id} 실패: {e}')
        with _hint_progress_lock:
            _hint_progress[unit_id]['done'] = True
            _hint_progress[unit_id]['running'] = False
            _hint_progress[unit_id]['error'] = str(e)
    finally:
        connection.close()


@teacher_required
@require_POST
def generate_hints_ajax(request, unit_id):
    """단원의 한글뜻 생성 시작 (백그라운드). body.force=true면 기존 한글뜻도 강제 재생성."""
    unit = get_object_or_404(WritingUnit, pk=unit_id)
    force = False
    try:
        if request.body:
            force = bool(json.loads(request.body).get('force', False))
    except (json.JSONDecodeError, ValueError, TypeError):
        force = False

    with _hint_progress_lock:
        existing = _hint_progress.get(unit_id)
        if existing and existing.get('running'):
            return JsonResponse({
                'started': False,
                'message': '이미 생성 중입니다.',
                **existing,
            })

        if force:
            # 모든 word_hints 초기화 → 다음 단계의 query가 전부 잡음
            WritingProblem.objects.filter(unit_id=unit_id).update(word_hints=[])

        _hint_progress[unit_id] = {
            'total': unit.problems.filter(word_hints=[]).count(),
            'completed': 0,
            'done': False,
            'running': True,
            'error': None,
        }

    if _hint_progress[unit_id]['total'] == 0:
        with _hint_progress_lock:
            _hint_progress[unit_id]['done'] = True
            _hint_progress[unit_id]['running'] = False
        return JsonResponse({'started': True, 'done': True, 'total': 0, 'completed': 0})

    # 백그라운드 스레드 시작
    t = threading.Thread(target=_run_hint_generation, args=(unit_id,), daemon=True)
    t.start()

    return JsonResponse({
        'started': True,
        'total': _hint_progress[unit_id]['total'],
        'completed': 0,
        'done': False,
    })


@teacher_required
@require_GET
def assignment_list(request, unit_id):
    """단원에 배정 가능한 학생 목록 + 현재 배정 상태 반환"""
    from django.contrib.auth import get_user_model
    User = get_user_model()
    unit = get_object_or_404(WritingUnit, pk=unit_id)

    assigned_ids = set(
        UnitAssignment.objects.filter(unit=unit).values_list('student_id', flat=True)
    )

    users = User.objects.filter(is_active=True).order_by('username')
    students = []
    for u in users:
        if is_teacher(u):
            continue
        name = (
            getattr(u, 'name', None)
            or getattr(u, 'full_name', None)
            or getattr(u, 'username', '')
            or ''
        )
        students.append({
            'id': u.id,
            'username': u.username,
            'name': name,
            'is_assigned': u.id in assigned_ids,
        })

    return JsonResponse(
        {'students': students, 'assigned_count': len(assigned_ids)},
        json_dumps_params={'ensure_ascii': False},
    )


@teacher_required
@require_POST
def assignment_update(request, unit_id):
    """학생 배정 일괄 갱신 — body: {student_ids:[..]} 의 학생만 배정 상태로 (나머지는 해제)"""
    unit = get_object_or_404(WritingUnit, pk=unit_id)
    try:
        data = json.loads(request.body)
        target_ids = {int(x) for x in data.get('student_ids', [])}
    except (json.JSONDecodeError, ValueError, TypeError):
        return HttpResponseBadRequest('Invalid')

    current_ids = set(
        UnitAssignment.objects.filter(unit=unit).values_list('student_id', flat=True)
    )
    to_add = target_ids - current_ids
    to_remove = current_ids - target_ids

    if to_add:
        UnitAssignment.objects.bulk_create([
            UnitAssignment(student_id=sid, unit=unit, assigned_by=request.user)
            for sid in to_add
        ], ignore_conflicts=True)
    if to_remove:
        UnitAssignment.objects.filter(unit=unit, student_id__in=to_remove).delete()

    return JsonResponse({
        'success': True,
        'assigned_count': UnitAssignment.objects.filter(unit=unit).count(),
        'added': len(to_add),
        'removed': len(to_remove),
    })


@teacher_required
@require_POST
def generate_hints_bulk_ajax(request):
    """체크한 N개 단원에서 word_hints 비어있는 문제 백그라운드 생성.
    body.force=true면 이미 한글뜻 있는 문제도 강제 재생성."""
    try:
        data = json.loads(request.body or '{}')
        unit_ids = [int(x) for x in data.get('unit_ids', [])]
        force = bool(data.get('force', False))
    except (json.JSONDecodeError, ValueError, TypeError):
        return HttpResponseBadRequest('Invalid unit_ids')

    started = []
    already_running = []
    already_done = []

    for uid in unit_ids:
        unit = WritingUnit.objects.filter(pk=uid).first()
        if not unit:
            continue
        if force:
            WritingProblem.objects.filter(unit_id=uid).update(word_hints=[])
        remaining = unit.problems.filter(word_hints=[]).count()
        if remaining == 0:
            already_done.append(uid)
            continue
        with _hint_progress_lock:
            existing = _hint_progress.get(uid)
            if existing and existing.get('running'):
                already_running.append(uid)
                continue
            _hint_progress[uid] = {
                'total': remaining, 'completed': 0,
                'done': False, 'running': True, 'error': None,
            }
        threading.Thread(target=_run_hint_generation, args=(uid,), daemon=True).start()
        started.append(uid)

    return JsonResponse({
        'started': len(started),
        'already_running': len(already_running),
        'already_done': len(already_done),
        'forced': force,
    })


DEFAULT_STUDENT_PASSWORD = '123456'


def _xlsx_response(workbook, filename):
    from io import BytesIO
    from django.http import HttpResponse
    buf = BytesIO()
    workbook.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _style_header_row(ws):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    bold = Font(bold=True)
    fill = PatternFill('solid', fgColor='E5E7EB')
    center = Alignment(horizontal='center', vertical='center')
    for cell in ws[1]:
        cell.font = bold
        cell.fill = fill
        cell.alignment = center


@teacher_required
def student_template_xlsx(request):
    """학생 일괄 등록용 빈 양식. 1열 색인 · 2열 ID · 3열 이름."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '학생 명단'
    ws.append(['색인', 'ID', '이름'])
    ws.append([1, 'dprimeedu1', '홍길동'])
    ws.append([2, 'dprimeedu2', '김철수'])
    ws.append([3, 'dprimeedu3', '이영희'])
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    _style_header_row(ws)
    return _xlsx_response(wb, 'students_template.xlsx')


@teacher_required
def writing_template_xlsx(request):
    """영작 단원 업로드용 빈 양식."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '영작 문제'
    ws.append(['색인', '영어', '한글'])
    ws.append([1, 'In under three years was the project finished.', '3년도 채 안 되어 프로젝트는 끝났다.'])
    ws.append([2, 'I had never seen such a beautiful sunset.', '나는 그렇게 아름다운 일몰을 본 적이 없었다.'])
    ws.append([3, 'You should always do your best.', '항상 최선을 다해야 한다.'])
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 35
    _style_header_row(ws)
    return _xlsx_response(wb, 'writing_unit_template.xlsx')


@teacher_required
def student_admin(request):
    """학생 관리 페이지 — 전체 학생 + 액션."""
    from django.contrib.auth import get_user_model
    User = get_user_model()
    qs = User.objects.exclude(is_staff=True).exclude(is_superuser=True).order_by('-date_joined')
    students = []
    for s in qs:
        if is_teacher(s):
            continue
        s.assigned_count = UnitAssignment.objects.filter(student=s).count()
        students.append(s)

    # 오늘 목표를 학생 행에 prefetch
    today = timezone.now().date()
    sid_list = [s.id for s in students]
    goals_today = {
        g.student_id: g
        for g in DailyStudyGoal.objects.filter(date=today, student_id__in=sid_list)
    }
    for s in students:
        s.today_goal = goals_today.get(s.id)

    return render(request, 'writing/student_list.html', {
        'students': students,
        'default_password': DEFAULT_STUDENT_PASSWORD,
        'today': today,
    })


@teacher_required
def student_upload(request):
    """엑셀로 학생 일괄 등록. 컬럼: 1열 ID(login_id), 2열 이름. 비번 공통."""
    if request.method != 'POST':
        return render(request, 'writing/student_upload.html', {
            'default_password': DEFAULT_STUDENT_PASSWORD,
        })

    f = request.FILES.get('excel_file')
    if not f:
        messages.error(request, '엑셀 파일을 선택해주세요.')
        return render(request, 'writing/student_upload.html', {
            'default_password': DEFAULT_STUDENT_PASSWORD,
        })
    if not f.name.lower().endswith(('.xlsx', '.xls')):
        messages.error(request, 'xlsx 또는 xls 파일만 가능합니다.')
        return render(request, 'writing/student_upload.html', {
            'default_password': DEFAULT_STUDENT_PASSWORD,
        })

    result = parse_students_excel(f)
    if not result['success']:
        for err in result['errors']:
            messages.error(request, err)
        return render(request, 'writing/student_upload.html', {
            'default_password': DEFAULT_STUDENT_PASSWORD,
        })

    from django.contrib.auth import get_user_model
    User = get_user_model()
    created = 0
    skipped = []
    for s in result['students']:
        login_id = s['login_id']
        name = s['name']
        if User.objects.filter(login_id=login_id).exists():
            skipped.append(login_id)
            continue
        try:
            user = User(
                login_id=login_id,
                username=name,
                email=None,
                member_type='user',
                is_active=True,
                is_approved=True,
                is_academy=False,
            )
            user.set_password(DEFAULT_STUDENT_PASSWORD)
            user.save()
            created += 1
        except Exception as e:
            skipped.append(f'{login_id} ({e})')

    parts = [f'{created}명 등록 완료', f'기본 비번: {DEFAULT_STUDENT_PASSWORD}']
    if skipped:
        sample = ', '.join(skipped[:5])
        more = '…' if len(skipped) > 5 else ''
        parts.append(f'중복/실패 {len(skipped)}건 skip ({sample}{more})')
    messages.success(request, ' · '.join(parts))
    return redirect('writing:student_admin')


@teacher_required
@require_POST
def student_action(request):
    """학생 일괄 액션 — action: approve|unapprove|activate|deactivate|delete"""
    raw_ids = request.POST.getlist('student_ids')
    action = request.POST.get('action')
    try:
        ids = [int(x) for x in raw_ids if x]
    except ValueError:
        ids = []

    if not ids or action not in ('approve', 'unapprove', 'activate', 'deactivate', 'delete'):
        messages.warning(request, '학생과 액션을 선택해주세요.')
        return redirect('writing:student_admin')

    from django.contrib.auth import get_user_model
    User = get_user_model()
    qs = User.objects.filter(pk__in=ids).exclude(is_staff=True).exclude(is_superuser=True)

    if action == 'approve':
        n = qs.update(is_approved=True)
        messages.success(request, f'{n}명 재원생 승인 완료.')
    elif action == 'unapprove':
        n = qs.update(is_approved=False)
        messages.success(request, f'{n}명 재원생 승인 취소.')
    elif action == 'activate':
        n = qs.update(is_active=True)
        messages.success(request, f'{n}명 계정 활성화.')
    elif action == 'deactivate':
        n = qs.update(is_active=False)
        messages.success(request, f'{n}명 계정 비활성화.')
    elif action == 'delete':
        n = qs.count()
        qs.delete()
        messages.success(request, f'{n}명 삭제 완료. (단원 배정·풀이 기록도 함께 삭제)')

    return redirect('writing:student_admin')


@teacher_required
@require_GET
def student_info_api(request, student_id):
    """학생 정보 + 풀이 통계 — 학생 정보 모달용."""
    from django.contrib.auth import get_user_model
    from django.db.models import Max, Avg, Sum
    User = get_user_model()
    s = get_object_or_404(
        User.objects.exclude(is_staff=True).exclude(is_superuser=True),
        pk=student_id,
    )
    assigned_count = UnitAssignment.objects.filter(student=s).count()
    finished = WritingSession.objects.filter(student=s, finished_at__isnull=False)
    session_count = finished.count()
    best_score = finished.aggregate(Max('total_score'))['total_score__max'] or 0
    profile = StudentProfile.objects.filter(student=s).first()
    total_xp = profile.total_xp if profile else 0
    level = scoring.compute_level(total_xp)

    return JsonResponse({
        'id': s.id,
        'username': s.username or '',
        'login_id': getattr(s, 'login_id', '') or '',
        'email': s.email or '',
        'is_active': s.is_active,
        'is_approved': bool(getattr(s, 'is_approved', False)),
        'date_joined': s.date_joined.strftime('%Y-%m-%d %H:%M') if s.date_joined else '',
        'assigned_count': assigned_count,
        'session_count': session_count,
        'best_score': best_score,
        'total_xp': total_xp,
        'level': level,
        'title': scoring.compute_title(level),
    }, json_dumps_params={'ensure_ascii': False})


@teacher_required
@require_GET
def student_assignments(request, student_id):
    """학생 1명의 배정 현황 + 전체 단원 — 학생 편집 모달용."""
    from django.contrib.auth import get_user_model
    User = get_user_model()
    student = get_object_or_404(
        User.objects.exclude(is_staff=True).exclude(is_superuser=True),
        pk=student_id,
    )
    assigned_ids = set(
        UnitAssignment.objects.filter(student=student).values_list('unit_id', flat=True)
    )
    units = []
    for u in WritingUnit.objects.filter(is_active=True).order_by('grade', 'publisher', 'title'):
        units.append({
            'id': u.id,
            'title': u.title,
            'grade': u.grade,
            'publisher': u.publisher,
            'problem_count': u.problem_count,
            'is_assigned': u.id in assigned_ids,
        })
    name = (
        getattr(student, 'name', None)
        or getattr(student, 'full_name', None)
        or getattr(student, 'username', '')
        or ''
    )
    return JsonResponse({
        'student': {
            'id': student.id,
            'username': student.username,
            'login_id': getattr(student, 'login_id', '') or '',
            'name': name,
        },
        'units': units,
        'assigned_count': len(assigned_ids),
    }, json_dumps_params={'ensure_ascii': False})


@teacher_required
@require_POST
def student_assignments_update(request, student_id):
    """학생의 배정을 body.unit_ids 로 통째 갱신 (체크 안 된 단원은 해제)."""
    from django.contrib.auth import get_user_model
    User = get_user_model()
    student = get_object_or_404(
        User.objects.exclude(is_staff=True).exclude(is_superuser=True),
        pk=student_id,
    )
    try:
        data = json.loads(request.body)
        target_ids = {int(x) for x in data.get('unit_ids', [])}
    except (json.JSONDecodeError, ValueError, TypeError):
        return HttpResponseBadRequest('Invalid')

    valid_ids = set(
        WritingUnit.objects.filter(pk__in=target_ids).values_list('id', flat=True)
    )
    current_ids = set(
        UnitAssignment.objects.filter(student=student).values_list('unit_id', flat=True)
    )
    to_add = valid_ids - current_ids
    to_remove = current_ids - valid_ids

    if to_add:
        UnitAssignment.objects.bulk_create([
            UnitAssignment(student=student, unit_id=uid, assigned_by=request.user)
            for uid in to_add
        ], ignore_conflicts=True)
    if to_remove:
        UnitAssignment.objects.filter(student=student, unit_id__in=to_remove).delete()

    return JsonResponse({
        'success': True,
        'assigned_count': UnitAssignment.objects.filter(student=student).count(),
        'added': len(to_add),
        'removed': len(to_remove),
    })


def _fmt_duration(secs: int) -> str:
    if not secs or secs < 60:
        return f'{secs or 0}초'
    h, rem = divmod(int(secs), 3600)
    m, _ = divmod(rem, 60)
    return f'{h}시간 {m}분' if h else f'{m}분'


@teacher_required
@require_GET
def student_report(request, student_id):
    """학생별 일일 학습 리포트 — 오늘(또는 ?date=YYYY-MM-DD) 기준."""
    from django.contrib.auth import get_user_model
    from datetime import timedelta as _td
    User = get_user_model()
    student = get_object_or_404(
        User.objects.exclude(is_staff=True).exclude(is_superuser=True),
        pk=student_id,
    )

    today = timezone.now().date()
    raw_date = (request.GET.get('date') or '').strip()
    try:
        target_date = datetime.strptime(raw_date, '%Y-%m-%d').date() if raw_date else today
    except ValueError:
        target_date = today

    # USE_TZ=False 라서 DB 저장값도 naive — 같은 naive 로 비교
    day_start = datetime.combine(target_date, datetime.min.time())
    day_end = day_start + _td(days=1)

    # 오늘 세션
    sessions_today = list(
        WritingSession.objects.filter(
            student=student,
            started_at__gte=day_start,
            started_at__lt=day_end,
        ).select_related('unit').order_by('started_at')
    )
    session_ids = [s.id for s in sessions_today]
    finished_today = [s for s in sessions_today if s.finished_at]

    total_study_seconds = sum(
        int((s.finished_at - s.started_at).total_seconds())
        for s in finished_today if s.started_at and s.finished_at
    )

    first_started = sessions_today[0].started_at if sessions_today else None
    last_ended = max((s.finished_at for s in finished_today), default=None)

    # 오늘 단어 시도
    attempts_today = list(
        WritingAttempt.objects.filter(session_id__in=session_ids)
        .select_related('problem__unit', 'session__unit')
    )

    total_attempts = len(attempts_today)
    correct_attempts = sum(1 for a in attempts_today if a.is_correct)
    word_accuracy = round(correct_attempts / total_attempts * 100, 1) if total_attempts else 0

    total_score_today = sum(s.total_score for s in sessions_today)
    perfect_sentences_today = sum(s.perfect_sentences for s in sessions_today)

    # 푼 문제 수 = 오늘 시도 흔적이 있는 distinct problem
    problems_touched_today = len({a.problem_id for a in attempts_today})

    # 오늘의 목표 + 달성률
    goal = DailyStudyGoal.objects.filter(student=student, date=target_date).first()
    study_minutes = total_study_seconds // 60
    goal_view = None
    if goal:
        def _pct(done, target):
            if not target:
                return None
            return min(100, round(done / target * 100))
        goal_view = {
            'obj': goal,
            'problems': {
                'target': goal.target_problems,
                'done': problems_touched_today,
                'pct': _pct(problems_touched_today, goal.target_problems),
            },
            'minutes': {
                'target': goal.target_minutes,
                'done': study_minutes,
                'pct': _pct(study_minutes, goal.target_minutes),
            },
            'sessions': {
                'target': goal.target_sessions,
                'done': len(finished_today),
                'pct': _pct(len(finished_today), goal.target_sessions),
            },
            'note': goal.note,
        }

    # 단원별 집계
    unit_stats_map = {}
    for s in sessions_today:
        st = unit_stats_map.setdefault(s.unit_id, {
            'unit': s.unit, 'sessions': 0, 'score': 0,
            'perfect': 0, 'attempts': 0, 'correct': 0,
        })
        st['sessions'] += 1
        st['score'] += s.total_score
        st['perfect'] += s.perfect_sentences
    for a in attempts_today:
        st = unit_stats_map.get(a.session.unit_id)
        if not st:
            continue
        st['attempts'] += 1
        if a.is_correct:
            st['correct'] += 1

    unit_stats = []
    for st in unit_stats_map.values():
        acc = round(st['correct'] / st['attempts'] * 100, 1) if st['attempts'] else 0
        unit_stats.append({**st, 'accuracy': acc})
    unit_stats.sort(key=lambda x: -x['score'])

    # 문장(problem) 단위 — 잘한/어려운
    by_problem = {}  # pid → list[attempts]
    for a in attempts_today:
        by_problem.setdefault(a.problem_id, []).append(a)

    good_sentences = []
    hard_sentences = []
    for pid, atts in by_problem.items():
        problem = atts[0].problem
        # 단어별 1차 시도 정보
        by_word = {}
        for a in atts:
            by_word.setdefault(a.word_index, []).append(a)
        all_perfect = True
        max_hint = 0
        max_try = 0
        for wi, ws in by_word.items():
            first = min(ws, key=lambda x: x.attempt_num)
            if not (first.attempt_num == 1 and first.hint_level == 0 and first.is_correct):
                all_perfect = False
            for a in ws:
                max_hint = max(max_hint, a.hint_level)
                max_try = max(max_try, a.attempt_num)
        entry = {
            'problem': problem,
            'unit_title': problem.unit.title,
            'max_hint': max_hint,
            'max_try': max_try,
        }
        if all_perfect:
            good_sentences.append(entry)
        elif max_hint >= 2 or max_try >= 3:
            hard_sentences.append(entry)

    # 어려운 문장은 시도/힌트 많은 순
    hard_sentences.sort(key=lambda x: (-x['max_hint'], -x['max_try']))

    # 학습 기록 없을 때 7일 fallback
    week_summary = None
    if not sessions_today:
        week_start = day_start - _td(days=7)
        recent = WritingSession.objects.filter(
            student=student,
            started_at__gte=week_start,
            started_at__lt=day_end,
        )
        week_summary = {
            'sessions': recent.count(),
            'finished': recent.filter(finished_at__isnull=False).count(),
            'score': sum(s.total_score for s in recent),
        }

    profile = StudentProfile.objects.filter(student=student).first()
    total_xp = profile.total_xp if profile else 0
    level = scoring.compute_level(total_xp)

    name = student.username or getattr(student, 'login_id', '') or '학생'

    GOOD_LIMIT = 10
    HARD_LIMIT = 10
    return render(request, 'writing/student_report.html', {
        'student': student,
        'student_name': name,
        'target_date': target_date,
        'today': today,
        'is_today': target_date == today,
        'prev_date': target_date - _td(days=1),
        'next_date': target_date + _td(days=1),
        'sessions_today': sessions_today,
        'session_count': len(sessions_today),
        'finished_count': len(finished_today),
        'study_duration': _fmt_duration(total_study_seconds),
        'study_seconds': total_study_seconds,
        'first_started': first_started,
        'last_ended': last_ended,
        'total_attempts': total_attempts,
        'correct_attempts': correct_attempts,
        'word_accuracy': word_accuracy,
        'total_score_today': total_score_today,
        'perfect_sentences_today': perfect_sentences_today,
        'unit_stats': unit_stats,
        'good_sentences': good_sentences[:GOOD_LIMIT],
        'good_count_total': len(good_sentences),
        'good_more': max(0, len(good_sentences) - GOOD_LIMIT),
        'hard_sentences': hard_sentences[:HARD_LIMIT],
        'hard_count_total': len(hard_sentences),
        'hard_more': max(0, len(hard_sentences) - HARD_LIMIT),
        'week_summary': week_summary,
        'total_xp': total_xp,
        'level': level,
        'goal': goal_view,
        'problems_touched_today': problems_touched_today,
    })


@teacher_required
@require_POST
def student_goal_update(request, student_id):
    """학생의 그날 학습 목표 저장 (POST). 모든 target=0이면 행 삭제."""
    from django.contrib.auth import get_user_model
    User = get_user_model()
    student = get_object_or_404(
        User.objects.exclude(is_staff=True).exclude(is_superuser=True),
        pk=student_id,
    )

    raw_date = (request.POST.get('date') or '').strip()
    try:
        target_date = datetime.strptime(raw_date, '%Y-%m-%d').date() if raw_date else timezone.now().date()
    except ValueError:
        return HttpResponseBadRequest('Invalid date')

    def _int(name):
        try:
            return max(0, int(request.POST.get(name) or 0))
        except (TypeError, ValueError):
            return 0

    tp = _int('target_problems')
    tm = _int('target_minutes')
    ts = _int('target_sessions')
    note = (request.POST.get('note') or '').strip()[:200]

    if tp == 0 and tm == 0 and ts == 0 and not note:
        DailyStudyGoal.objects.filter(student=student, date=target_date).delete()
        messages.success(request, f'{target_date} 목표 삭제됨.')
    else:
        DailyStudyGoal.objects.update_or_create(
            student=student, date=target_date,
            defaults={
                'target_problems': tp,
                'target_minutes': tm,
                'target_sessions': ts,
                'note': note,
                'set_by': request.user,
            },
        )
        messages.success(request, f'{target_date} 목표 저장됨.')

    from django.urls import reverse
    return redirect(f"{reverse('writing:student_report', args=[student.id])}?date={target_date}")


@teacher_required
@require_GET
def student_goal_api(request, student_id):
    """학생 1명의 특정 날짜 목표 조회 (모달 prefill용)."""
    from django.contrib.auth import get_user_model
    User = get_user_model()
    student = get_object_or_404(
        User.objects.exclude(is_staff=True).exclude(is_superuser=True),
        pk=student_id,
    )
    raw_date = (request.GET.get('date') or '').strip()
    try:
        target_date = datetime.strptime(raw_date, '%Y-%m-%d').date() if raw_date else timezone.now().date()
    except ValueError:
        return HttpResponseBadRequest('Invalid date')

    g = DailyStudyGoal.objects.filter(student=student, date=target_date).first()
    return JsonResponse({
        'student_id': student.id,
        'date': target_date.strftime('%Y-%m-%d'),
        'target_problems': g.target_problems if g else 0,
        'target_minutes': g.target_minutes if g else 0,
        'target_sessions': g.target_sessions if g else 0,
        'note': g.note if g else '',
    }, json_dumps_params={'ensure_ascii': False})


@teacher_required
@require_POST
def student_goal_save_api(request):
    """단일/일괄 목표 저장 API.

    Body: { student_ids: [int], date: 'YYYY-MM-DD', target_problems, target_minutes,
            target_sessions, note }
    모든 target=0 + note 비면 그 학생들의 해당 날짜 목표 삭제.
    """
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return HttpResponseBadRequest('Invalid JSON')

    raw_ids = data.get('student_ids') or []
    try:
        student_ids = [int(x) for x in raw_ids]
    except (TypeError, ValueError):
        return HttpResponseBadRequest('Invalid student_ids')
    if not student_ids:
        return HttpResponseBadRequest('No students')

    raw_date = (data.get('date') or '').strip()
    try:
        target_date = datetime.strptime(raw_date, '%Y-%m-%d').date() if raw_date else timezone.now().date()
    except ValueError:
        return HttpResponseBadRequest('Invalid date')

    def _int(name):
        try:
            return max(0, int(data.get(name) or 0))
        except (TypeError, ValueError):
            return 0

    tp = _int('target_problems')
    tm = _int('target_minutes')
    ts = _int('target_sessions')
    note = (data.get('note') or '').strip()[:200]

    from django.contrib.auth import get_user_model
    User = get_user_model()
    valid_ids = list(
        User.objects.filter(pk__in=student_ids)
        .exclude(is_staff=True).exclude(is_superuser=True)
        .values_list('id', flat=True)
    )

    is_clear = (tp == 0 and tm == 0 and ts == 0 and not note)
    if is_clear:
        deleted = DailyStudyGoal.objects.filter(
            student_id__in=valid_ids, date=target_date,
        ).delete()[0]
        return JsonResponse({
            'success': True, 'action': 'delete',
            'affected': deleted, 'student_ids': valid_ids,
        })

    saved = 0
    for sid in valid_ids:
        DailyStudyGoal.objects.update_or_create(
            student_id=sid, date=target_date,
            defaults={
                'target_problems': tp,
                'target_minutes': tm,
                'target_sessions': ts,
                'note': note,
                'set_by': request.user,
            },
        )
        saved += 1
    return JsonResponse({
        'success': True, 'action': 'save',
        'affected': saved, 'student_ids': valid_ids,
        'target_problems': tp, 'target_minutes': tm,
        'target_sessions': ts, 'note': note,
        'date': target_date.strftime('%Y-%m-%d'),
    }, json_dumps_params={'ensure_ascii': False})


@teacher_required
@require_GET
def student_list_api(request):
    """전체 학생 목록 (단원 무관) — 일괄 배정/해제 모달용."""
    from django.contrib.auth import get_user_model
    User = get_user_model()
    users = User.objects.filter(is_active=True).order_by('username')
    students = []
    for u in users:
        if is_teacher(u):
            continue
        name = (
            getattr(u, 'name', None)
            or getattr(u, 'full_name', None)
            or getattr(u, 'username', '')
            or ''
        )
        students.append({'id': u.id, 'username': u.username, 'name': name})
    return JsonResponse({'students': students}, json_dumps_params={'ensure_ascii': False})


@teacher_required
@require_POST
def assignments_bulk_update(request):
    """체크한 N개 단원에 학생 추가 또는 단원에서 학생 제거.
    body: { action: 'add'|'remove', unit_ids: [...], student_ids: [...] }
    """
    try:
        data = json.loads(request.body)
        action = data.get('action')
        unit_ids = [int(x) for x in data.get('unit_ids', [])]
        student_ids = [int(x) for x in data.get('student_ids', [])]
    except (json.JSONDecodeError, ValueError, TypeError):
        return HttpResponseBadRequest('Invalid body')

    if action not in ('add', 'remove'):
        return HttpResponseBadRequest('action must be add or remove')
    if not unit_ids or not student_ids:
        return HttpResponseBadRequest('Empty unit_ids or student_ids')

    valid_unit_ids = list(
        WritingUnit.objects.filter(pk__in=unit_ids).values_list('id', flat=True)
    )

    if action == 'add':
        UnitAssignment.objects.bulk_create(
            [
                UnitAssignment(student_id=sid, unit_id=uid, assigned_by=request.user)
                for uid in valid_unit_ids
                for sid in student_ids
            ],
            ignore_conflicts=True,
        )
        return JsonResponse({
            'success': True, 'action': 'add',
            'units': len(valid_unit_ids), 'students': len(student_ids),
        })

    deleted, _ = UnitAssignment.objects.filter(
        unit_id__in=valid_unit_ids, student_id__in=student_ids,
    ).delete()
    return JsonResponse({
        'success': True, 'action': 'remove',
        'units': len(valid_unit_ids), 'students': len(student_ids),
        'removed_rows': deleted,
    })


@teacher_required
@require_GET
def generate_hints_status(request, unit_id):
    """진행 상태 폴링.
    응답에:
      - completed/total: 이번 회차 처리 진행도 (기존 detail.html 폴링용 — 호환 유지)
      - has_hints/unit_total: 단원 전체 기준 현재 보유 한글뜻 수 (unit_list 표 갱신용)
      - running, done
    """
    unit = get_object_or_404(WritingUnit, pk=unit_id)
    has_hints = unit.problems.exclude(word_hints=[]).count()
    unit_total = unit.problems.count()

    with _hint_progress_lock:
        state = _hint_progress.get(unit_id)
        if state:
            payload = dict(state)
        else:
            payload = {'total': unit_total, 'completed': has_hints, 'done': True, 'running': False}

    payload['has_hints'] = has_hints
    payload['unit_total'] = unit_total
    return JsonResponse(payload)


# ─────────────────────────────────────────────
# 버그 신고
# ─────────────────────────────────────────────

@login_required
@require_POST
def bug_report_create(request):
    """학생/사용자의 버그 신고 접수.
    body: {session_id?, problem_id?, description?, screen_state?, url?}
    """
    try:
        data = json.loads(request.body or '{}')
    except json.JSONDecodeError:
        return HttpResponseBadRequest('Invalid JSON')

    session = None
    problem = None
    unit = None
    sid = data.get('session_id')
    pid = data.get('problem_id')
    if sid:
        session = WritingSession.objects.filter(pk=sid).first()
        if session:
            unit = session.unit
    if pid:
        problem = WritingProblem.objects.filter(pk=pid).first()
        if problem and not unit:
            unit = problem.unit

    description = str(data.get('description', '')).strip()

    # XP 보상 결정
    # - 내용이 비어 있으면 보상 X
    # - 이미 같은 학생-같은 문제에 XP 받은 신고가 있으면 보상 X (다중 신고로 farming 방지)
    # - 최근 누적 롤백이 BUG_REPORT_ROLLBACK_LIMIT 이상이면 보상 X (남용 학생)
    xp_awarded = 0
    award_reason = 'empty'
    if description and problem and unit:
        prior_paid = BugReport.objects.filter(
            student=request.user, problem=problem, xp_awarded__gt=0,
        ).exists()
        rollback_count = BugReport.objects.filter(
            student=request.user, xp_rolled_back=True,
        ).count()
        if prior_paid:
            award_reason = 'duplicate_problem'
        elif rollback_count >= scoring.BUG_REPORT_ROLLBACK_LIMIT:
            award_reason = 'rollback_limit'
        else:
            sul = level_service.get_or_create_unit_level(request.user, unit)
            mult = level_service.xp_multiplier(sul.level)
            perfect_xp = scoring.calc_perfect_unit_xp(unit, level_multiplier=mult)
            xp_awarded = perfect_xp * scoring.BUG_REPORT_REWARD_MULTIPLIER
            award_reason = 'paid'

    report = BugReport.objects.create(
        student=request.user,
        session=session,
        problem=problem,
        unit=unit,
        url=str(data.get('url', ''))[:500],
        description=description[:5000],
        screen_state=data.get('screen_state') if isinstance(data.get('screen_state'), dict) else {},
        xp_awarded=xp_awarded,
    )

    if xp_awarded > 0:
        profile = get_or_create_profile(request.user)
        profile.total_xp += xp_awarded
        profile.save(update_fields=['total_xp'])

    return JsonResponse({
        'success': True, 'id': report.id,
        'xp_awarded': xp_awarded,
        'award_reason': award_reason,
        'rollback_limit': scoring.BUG_REPORT_ROLLBACK_LIMIT,
    })


@teacher_required
def bug_report_list(request):
    """관리자 — 버그 신고 목록"""
    status_filter = request.GET.get('status', '')
    qs = BugReport.objects.select_related('student', 'unit', 'problem').order_by('-created_at')
    if status_filter:
        qs = qs.filter(status=status_filter)
    return render(request, 'writing/bug_report_list.html', {
        'reports': qs[:200],
        'status_filter': status_filter,
        'status_choices': BugReport.STATUS_CHOICES,
        'pending_count': BugReport.objects.filter(status='pending').count(),
    })


@teacher_required
def bug_report_detail(request, report_id):
    """관리자 — 버그 신고 1건 상세 + 상태/메모 갱신"""
    report = get_object_or_404(BugReport, pk=report_id)
    if request.method == 'POST':
        new_status = request.POST.get('status')
        note = request.POST.get('admin_note', '')
        if new_status in dict(BugReport.STATUS_CHOICES):
            report.status = new_status
        report.admin_note = note
        report.save(update_fields=['status', 'admin_note', 'updated_at'])
        messages.success(request, '저장되었습니다.')
        return redirect('writing:bug_report_detail', report_id=report.id)
    # 학생의 누적 롤백 신고 수 (5회 이상이면 이후 신고에 XP 지급 X)
    student_rollback_count = 0
    if report.student:
        student_rollback_count = BugReport.objects.filter(
            student=report.student, xp_rolled_back=True,
        ).count()
    return render(request, 'writing/bug_report_detail.html', {
        'report': report,
        'status_choices': BugReport.STATUS_CHOICES,
        'student_rollback_count': student_rollback_count,
        'rollback_limit': scoring.BUG_REPORT_ROLLBACK_LIMIT,
    })


@teacher_required
@require_POST
def bug_report_rollback(request, report_id):
    """관리자가 '이상 없음' 판단 — 학생 XP 회수.

    같은 학생의 누적 롤백이 BUG_REPORT_ROLLBACK_LIMIT 이상이면
    그 학생은 추가 신고에 XP 지급 안 됨 (bug_report_create의 가드).
    """
    report = get_object_or_404(BugReport, pk=report_id)
    if report.xp_rolled_back:
        messages.info(request, '이미 회수된 신고입니다.')
    elif report.xp_awarded <= 0:
        messages.info(request, '이 신고는 지급된 XP가 없습니다.')
    elif not report.student:
        messages.error(request, '학생 정보가 없어 회수할 수 없습니다.')
    else:
        profile, _ = StudentProfile.objects.get_or_create(student=report.student)
        profile.total_xp = max(0, profile.total_xp - report.xp_awarded)
        profile.save(update_fields=['total_xp'])
        report.xp_rolled_back = True
        report.xp_rolled_back_at = timezone.now()
        report.status = 'dismissed'
        report.save(update_fields=[
            'xp_rolled_back', 'xp_rolled_back_at', 'status', 'updated_at',
        ])
        messages.success(
            request,
            f'{report.student.username}에서 {report.xp_awarded} XP 회수 완료',
        )
    return redirect('writing:bug_report_detail', report_id=report.id)
