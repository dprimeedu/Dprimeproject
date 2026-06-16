"""모의고사 회차·문항별 단어 DB 적재.

소스: '고1 - 고3 전체모의고사 단어 DB.xlsx' (시트 고1/고2/고3),
컬럼 = 번호 / 단어 / 의미 / 년도 / 학년 / 월 / 검색 / 유형 / 번호.

사용 예
  # 로컬: 엑셀에서 직접 적재 (+ 운영 파이프용 JSONL 덤프)
  python manage.py import_mock_vocab --xlsx "…/고1 - 고3 전체모의고사 단어 DB.xlsx" --jsonl-out mock_vocab.jsonl

  # 운영(컨테이너): JSONL을 stdin으로
  ssh dprime-nas 'sudo docker exec -i django …/python …/manage.py import_mock_vocab --stdin' < mock_vocab.jsonl

옵션
  --only 3:2011:9,3:2011:10   학년:년도:월 회차만 (테스트용)
  --replace                   적재 전 (해당 범위) 전체 삭제 후 새로 넣기
JSONL 한 줄 = {"grade":3,"year":2011,"month":9,"number":18,"word":"genius","meaning":"천재"}
"""
import json
import sys

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from vocab.models import MockVocab


def norm_key(word):
    return ' '.join((word or '').split()).lower()[:200]


SHEET_GRADE = {'고1': 1, '고2': 2, '고3': 3}


def _rows_from_xlsx(path, only):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sn, grade in SHEET_GRADE.items():
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        for r in ws.iter_rows(min_row=2, values_only=True):
            # 번호, 단어, 의미, 년도, 학년, 월, ...
            number, word, meaning, year, _g, month = (r + (None,) * 6)[:6]
            if not word or meaning in (None, ''):
                continue
            try:
                number, year, month = int(number), int(year), int(month)
            except (TypeError, ValueError):
                continue
            if only and (grade, year, month) not in only:
                continue
            yield {'grade': grade, 'year': year, 'month': month, 'number': number,
                   'word': str(word).strip()[:200], 'meaning': str(meaning).strip()}


def _rows_from_stdin(only):
    try:
        sys.stdin.reconfigure(encoding='utf-8')   # Windows 기본 cp949 stdin 방지
    except Exception:
        pass
    for line in sys.stdin:
        line = line.strip()
        if not line:
            continue
        try:
            d = json.loads(line)
            rec = {'grade': int(d['grade']), 'year': int(d['year']), 'month': int(d['month']),
                   'number': int(d['number']), 'word': str(d['word']).strip()[:200],
                   'meaning': str(d['meaning']).strip()}
        except (ValueError, KeyError, TypeError):
            continue
        if not rec['word'] or not rec['meaning']:
            continue
        if only and (rec['grade'], rec['year'], rec['month']) not in only:
            continue
        yield rec


class Command(BaseCommand):
    help = '모의고사 회차·문항별 단어 DB 적재 (엑셀 또는 JSONL stdin)'

    def add_arguments(self, parser):
        parser.add_argument('--xlsx', help='엑셀 경로')
        parser.add_argument('--stdin', action='store_true', help='JSONL을 stdin에서 읽기')
        parser.add_argument('--jsonl-out', help='읽은 행을 JSONL로 덤프(운영 파이프용)')
        parser.add_argument('--only', default='', help='학년:년도:월 회차 콤마목록 (예 3:2011:9,3:2011:10)')
        parser.add_argument('--replace', action='store_true', help='대상 범위 삭제 후 적재')

    def handle(self, *args, **opts):
        only = set()
        for tok in (opts['only'] or '').split(','):
            tok = tok.strip()
            if not tok:
                continue
            try:
                g, y, m = (int(x) for x in tok.split(':'))
                only.add((g, y, m))
            except ValueError:
                raise CommandError(f'--only 형식 오류: {tok} (학년:년도:월)')

        if opts['xlsx']:
            rows_iter = _rows_from_xlsx(opts['xlsx'], only)
        elif opts['stdin']:
            rows_iter = _rows_from_stdin(only)
        else:
            raise CommandError('--xlsx 또는 --stdin 중 하나가 필요합니다.')

        dump = open(opts['jsonl_out'], 'w', encoding='utf-8') if opts.get('jsonl_out') else None

        # 회차 내 word_key 중복은 마지막 값으로(맥락 단어는 보통 1개). 메모리 적재 후 일괄 upsert.
        bucket = {}
        seen_scopes = set()
        total_in = 0
        for rec in rows_iter:
            total_in += 1
            if dump:
                dump.write(json.dumps(rec, ensure_ascii=False) + '\n')
            key = (rec['grade'], rec['year'], rec['month'], rec['number'], norm_key(rec['word']))
            bucket[key] = rec
            seen_scopes.add((rec['grade'], rec['year'], rec['month']))
        if dump:
            dump.close()

        if opts['replace'] and seen_scopes:
            with transaction.atomic():
                for (g, y, m) in seen_scopes:
                    MockVocab.objects.filter(grade=g, year=y, month=m).delete()

        objs = [
            MockVocab(grade=k[0], year=k[1], month=k[2], number=k[3],
                      word=rec['word'], word_key=k[4], meaning=rec['meaning'])
            for k, rec in bucket.items()
        ]
        created = 0
        with transaction.atomic():
            for i in range(0, len(objs), 2000):
                chunk = objs[i:i + 2000]
                MockVocab.objects.bulk_create(chunk, ignore_conflicts=True, batch_size=1000)
                created += len(chunk)

        self.stdout.write(self.style.SUCCESS(
            f'읽은 행 {total_in} · 고유 {len(objs)} · 적재(시도) {created} · 회차 {len(seen_scopes)}개'
            f"{' · replace' if opts['replace'] else ''}"))
        self.stdout.write(f'현재 MockVocab 총 {MockVocab.objects.count()}행')
