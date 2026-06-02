"""
통합본 xlsx '내신단어' 시트 → VocabUnit + VocabWord 가져오기.

내신단어 시트 형식: [색인, 단어, 해석, 단원, 출처] (0행 헤더, 1행부터 데이터)
예: Z:\\home\\Drive\\교재폴더\\내신\\...\\동백고1 1학기 기말고사 통합.xlsx

사용:
    python manage.py import_vocab "동백고1 1학기 기말고사 통합.xlsx"
    python manage.py import_vocab "...xlsx" --title "동백고1 1학기 기말고사" --grade 고1 --school 동백고1 --exam "1학기 기말고사"
    python manage.py import_vocab "...xlsx" --replace   # 기존 단어 전부 지우고 다시
    python manage.py import_vocab "...xlsx" --dry-run    # 저장 없이 미리보기
"""
import os
import re

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

import openpyxl

from vocab.models import VocabUnit, VocabWord

SHEET_NAME = '내신단어'
# 헤더명 → 내부 키 (헤더 순서가 달라도 이름으로 매핑)
HEADER_MAP = {
    '색인': 'index',
    '단어': 'word',
    '해석': 'meaning',
    '단원': 'sub_unit',
    '출처': 'source',
}


def _title_from_filename(path: str) -> str:
    """'동백고1 1학기 기말고사 통합.xlsx' → '동백고1 1학기 기말고사'."""
    name = os.path.splitext(os.path.basename(path))[0]
    # 끝의 ' 통합' / '통합' 및 괄호 메모(병합전백업 등) 제거
    name = re.sub(r'\s*\([^)]*\)\s*$', '', name)
    name = re.sub(r'\s*통합\s*$', '', name).strip()
    return name or os.path.basename(path)


class Command(BaseCommand):
    help = "통합본 xlsx의 '내신단어' 시트를 읽어 VocabUnit + VocabWord 생성/갱신"

    def add_arguments(self, parser):
        parser.add_argument('path', help='통합본 xlsx 파일 경로')
        parser.add_argument('--title', default=None, help='단원명 (기본: 파일명에서 추출)')
        parser.add_argument('--grade', default='기타', help='학년 (예: 고1)')
        parser.add_argument('--school', default='', help='학교')
        parser.add_argument('--exam', default='', help='시험')
        parser.add_argument('--sheet', default=SHEET_NAME, help=f'시트명 (기본: {SHEET_NAME})')
        parser.add_argument('--replace', action='store_true', help='기존 단어를 전부 지우고 다시 넣기')
        parser.add_argument('--dry-run', action='store_true', help='DB 저장 없이 미리보기만')

    def handle(self, *args, **opts):
        # Windows 콘솔(cp949)에서도 한글/특수문자 출력이 깨지지 않게 utf-8 강제
        try:
            self.stdout._out.reconfigure(encoding='utf-8')
        except Exception:
            pass

        path = opts['path']
        if not os.path.isfile(path):
            raise CommandError(f'파일을 찾을 수 없습니다: {path}')

        title = opts['title'] or _title_from_filename(path)
        sheet = opts['sheet']

        rows = self._read_sheet(path, sheet)
        if not rows:
            raise CommandError(f"'{sheet}' 시트에서 데이터를 찾지 못했습니다.")

        self.stdout.write(f"단원: '{title}'  (학년 {opts['grade']})  · 단어 {len(rows)}개")
        # 미리보기 3개
        for r in rows[:3]:
            self.stdout.write(f"  · [{r['index']}] {r['word']} — {r['meaning']}  ({r['sub_unit']}/{r['source']})")

        if opts['dry_run']:
            self.stdout.write(self.style.WARNING('[dry-run] 저장하지 않고 종료합니다.'))
            return

        with transaction.atomic():
            unit, created = VocabUnit.objects.get_or_create(
                title=title,
                defaults={
                    'grade': opts['grade'],
                    'school': opts['school'],
                    'exam': opts['exam'],
                },
            )
            if not created:
                # 메타 갱신 (명시된 값만)
                changed = []
                if opts['grade'] != '기타' and unit.grade != opts['grade']:
                    unit.grade = opts['grade']; changed.append('grade')
                if opts['school'] and unit.school != opts['school']:
                    unit.school = opts['school']; changed.append('school')
                if opts['exam'] and unit.exam != opts['exam']:
                    unit.exam = opts['exam']; changed.append('exam')
                if changed:
                    unit.save(update_fields=changed)

            if opts['replace']:
                deleted, _ = VocabWord.objects.filter(unit=unit).delete()
                if deleted:
                    self.stdout.write(self.style.WARNING(f'기존 단어 {deleted}개 삭제'))

            existing = {w.index: w for w in VocabWord.objects.filter(unit=unit)}
            to_create, to_update = [], []
            for r in rows:
                w = existing.get(r['index'])
                if w is None:
                    to_create.append(VocabWord(
                        unit=unit, index=r['index'], word=r['word'],
                        meaning=r['meaning'], sub_unit=r['sub_unit'], source=r['source'],
                    ))
                else:
                    w.word, w.meaning, w.sub_unit, w.source = r['word'], r['meaning'], r['sub_unit'], r['source']
                    to_update.append(w)

            if to_create:
                VocabWord.objects.bulk_create(to_create, batch_size=500)
            if to_update:
                VocabWord.objects.bulk_update(
                    to_update, ['word', 'meaning', 'sub_unit', 'source'], batch_size=500,
                )

        verb = '생성' if created else '갱신'
        self.stdout.write(self.style.SUCCESS(
            f"완료 — 단원 {verb}: '{unit.title}' (id={unit.id}) · 신규 {len(to_create)} / 수정 {len(to_update)}"
        ))

    def _read_sheet(self, path, sheet):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if sheet not in wb.sheetnames:
            wb.close()
            raise CommandError(
                f"'{sheet}' 시트가 없습니다. 사용 가능한 시트: {wb.sheetnames}"
            )
        ws = wb[sheet]

        it = ws.iter_rows(values_only=True)
        try:
            header = next(it)
        except StopIteration:
            wb.close()
            return []

        # 헤더명 → 컬럼 인덱스
        col = {}
        for i, h in enumerate(header):
            key = HEADER_MAP.get(str(h).strip() if h is not None else '')
            if key:
                col[key] = i
        missing = {'index', 'word', 'meaning'} - col.keys()
        if missing:
            wb.close()
            raise CommandError(
                f"헤더에서 필수 컬럼을 찾지 못했습니다: {missing}. 실제 헤더: {header}"
            )

        def cell(row, key):
            i = col.get(key)
            if i is None or i >= len(row):
                return ''
            v = row[i]
            return '' if v is None else str(v).strip()

        rows = []
        for raw in it:
            word = cell(raw, 'word')
            if not word:
                continue  # 빈 행 스킵
            idx_raw = cell(raw, 'index')
            try:
                index = int(float(idx_raw)) if idx_raw else len(rows) + 1
            except ValueError:
                index = len(rows) + 1
            rows.append({
                'index': index,
                'word': word,
                'meaning': cell(raw, 'meaning'),
                'sub_unit': cell(raw, 'sub_unit'),
                'source': cell(raw, 'source'),
            })
        wb.close()
        return rows
