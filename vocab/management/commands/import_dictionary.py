"""전체 단어장 모음 → 사전 DB(DictionaryEntry) 적재.

원본: '단어장 전체 영영전체모음.xlsm' 시트 '출제지'
컬럼: A=번호, B=영어, C=한글, D/E=영영(무시). B(영어)+C(한글)만 사용.

로컬에서 xlsx로 적재:
    python manage.py import_dictionary --replace --export dict_export.json
운영(NAS, xlsm 접근 불가)에서 JSON으로 적재:
    python manage.py import_dictionary --json /tmp/dict_export.json --replace
"""
import json
import os
import re

from django.core.management.base import BaseCommand, CommandError

from vocab.models import DictionaryEntry

DEFAULT_XLSX = r'Z:\home\Drive\교재폴더\어휘\(전체모음)\단어장 전체 영영전체모음.xlsm'

_HANGUL = re.compile(r'[가-힣]')
_WORD_RE = re.compile(r"^[A-Za-z][A-Za-z'.\-/() ]{0,39}$")


def _is_word(v):
    """영어 단어(짧은 형태)인지 — 예문/문장 컬럼을 배제."""
    if v is None:
        return False
    s = str(v).strip()
    if not s or _HANGUL.search(s) or not _WORD_RE.match(s):
        return False
    return len(s.split()) <= 4


def _has_kor(v):
    return v is not None and bool(_HANGUL.search(str(v)))


def _extract_pairs(path, max_cols=12):
    """단어장 xlsx에서 (영어, 한글) 추출 — 영어/한글 컬럼 자동 감지."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    ncol = min(max_cols, max((len(r) for r in rows[:50]), default=0))
    if not ncol:
        return []
    eng_score = [0] * ncol
    kor_score = [0] * ncol
    for r in rows[:80]:
        for ci in range(min(ncol, len(r))):
            if _is_word(r[ci]):
                eng_score[ci] += 1
            if _has_kor(r[ci]):
                kor_score[ci] += 1
    eng_col = max(range(ncol), key=lambda c: eng_score[c])
    kor_col = max(range(ncol), key=lambda c: kor_score[c])
    if eng_score[eng_col] < 3 or kor_score[kor_col] < 3:
        return []   # 단어장 형식 아님
    pairs = []
    for r in rows:
        if eng_col >= len(r) or kor_col >= len(r):
            continue
        w, m = r[eng_col], r[kor_col]
        if not _is_word(w) or not _has_kor(m):
            continue
        pairs.append((str(w).strip(), str(m).strip()))
    return pairs


class Command(BaseCommand):
    help = '전체 단어장 모음(영어 B열 / 한글 C열)을 사전 DB로 적재'

    def add_arguments(self, parser):
        parser.add_argument('--xlsx', default=None, help=f'엑셀 경로 (기본: {DEFAULT_XLSX})')
        parser.add_argument('--sheet', default='출제지', help='시트명 (기본: 출제지)')
        parser.add_argument('--json', default=None, help='사전 JSON([{word,meaning}]) 적재 (운영 전송용)')
        parser.add_argument('--from-vocab', action='store_true',
                            help='엑셀 대신 기존 VocabWord(내신단어 등) 전체를 사전에 합치기(없는 것만)')
        parser.add_argument('--folder', default=None,
                            help='폴더 안 단어장 xlsx 전체를 영어/한글 자동감지로 사전에 합치기(없는 것만)')
        parser.add_argument('--recursive', action='store_true', help='--folder 하위 폴더까지')
        parser.add_argument('--exclude', default='내신단어,파생어휘,내신영작',
                            help='--folder에서 제외할 파일명 포함어(쉼표구분). 기본: 내신단어,파생어휘,내신영작')
        parser.add_argument('--export', default=None, help='적재 후 JSON 익스포트 경로')
        parser.add_argument('--replace', action='store_true', help='기존 사전 전체 삭제 후 적재')
        parser.add_argument('--dry-run', action='store_true', help='적재 없이 집계만')

    def handle(self, *args, **o):
        # ── 기존 VocabWord 전체를 사전에 합치기(backfill) ──
        if o['from_vocab']:
            from vocab.models import VocabWord
            from vocab.services import sync_pairs_to_dictionary
            pairs = list(
                VocabWord.objects.exclude(meaning='')
                .values_list('word', 'meaning')
            )
            self.stdout.write(f'VocabWord {len(pairs)}개 → 통합 사전 합치는 중...')
            if o['dry_run']:
                self.stdout.write('dry-run — 적재 안 함')
                return
            added = sync_pairs_to_dictionary(pairs)
            self.stdout.write(self.style.SUCCESS(
                f'신규 {added}개 추가 (사전 총 {DictionaryEntry.objects.count()}개)'))
            return

        # ── 원본 읽기 ──
        pairs = []  # [(word, meaning)]
        if o['json']:
            if not os.path.exists(o['json']):
                raise CommandError(f'JSON 파일 없음: {o["json"]}')
            with open(o['json'], encoding='utf-8') as f:
                data = json.load(f)
            for d in data:
                pairs.append((str(d.get('word', '')), str(d.get('meaning', ''))))
        elif o['folder']:
            folder = o['folder']
            if not os.path.isdir(folder):
                raise CommandError(f'폴더 없음: {folder}')
            excludes = [x.strip() for x in (o['exclude'] or '').split(',') if x.strip()]
            files = []
            if o['recursive']:
                for root, _dirs, fs in os.walk(folder):
                    for f in fs:
                        files.append(os.path.join(root, f))
            else:
                files = [os.path.join(folder, f) for f in os.listdir(folder)]
            files = [
                f for f in files
                if f.lower().endswith(('.xlsx', '.xlsm')) and os.path.isfile(f)
                and not any(x in os.path.basename(f) for x in excludes)
            ]
            files.sort()
            self.stdout.write(f'단어장 파일 {len(files)}개 스캔 (제외어: {excludes})')
            ok = 0
            for fp in files:
                try:
                    fpairs = _extract_pairs(fp)
                except Exception as e:
                    self.stdout.write(self.style.WARNING(f'  - {os.path.basename(fp)} 읽기 실패: {e}'))
                    continue
                if fpairs:
                    ok += 1
                    pairs.extend(fpairs)
                    self.stdout.write(f'  · {os.path.basename(fp)[:40]:40s} {len(fpairs)}개')
                else:
                    self.stdout.write(self.style.WARNING(f'  - {os.path.basename(fp)} 단어 추출 0 (형식 불일치 → 건너뜀)'))
            self.stdout.write(f'단어장 {ok}개에서 {len(pairs)}행 추출')
        else:
            try:
                import openpyxl
            except ImportError:
                raise CommandError('openpyxl 필요')
            path = o['xlsx'] or DEFAULT_XLSX
            if not os.path.exists(path):
                raise CommandError(f'엑셀 없음: {path}')
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            if o['sheet'] not in wb.sheetnames:
                raise CommandError(f'시트 없음: {o["sheet"]} (있는 시트: {wb.sheetnames})')
            ws = wb[o['sheet']]
            first = True
            for row in ws.iter_rows(values_only=True):
                if first:                       # 헤더(번호/영어/한글/...) 스킵
                    first = False
                    continue
                word = '' if (len(row) < 2 or row[1] is None) else str(row[1]).strip()
                meaning = '' if (len(row) < 3 or row[2] is None) else str(row[2]).strip()
                if word and meaning:
                    pairs.append((word, meaning))

        # ── 정리: 소문자 key 기준 중복 제거(첫 항목 유지) ──
        seen = set()
        clean = []  # [(word, key, meaning)]
        for word, meaning in pairs:
            word = word.strip()[:255]
            meaning = meaning.strip()
            key = word.lower()
            if not (word and meaning and key) or key in seen:
                continue
            seen.add(key)
            clean.append((word, key, meaning))

        self.stdout.write(f'읽음 {len(pairs)}행 → 유니크 단어 {len(clean)}개')

        # ── 익스포트(운영 전송용) ──
        if o['export']:
            with open(o['export'], 'w', encoding='utf-8') as f:
                json.dump([{'word': w, 'meaning': m} for w, k, m in clean],
                          f, ensure_ascii=False)
            self.stdout.write(self.style.SUCCESS(f'익스포트 완료: {o["export"]} ({len(clean)}개)'))

        if o['dry_run']:
            self.stdout.write('dry-run — 적재 안 함')
            return

        # ── 적재 ──
        if o['replace']:
            n = DictionaryEntry.objects.count()
            DictionaryEntry.objects.all().delete()
            self.stdout.write(f'기존 {n}개 삭제')

        existing = set()
        if not o['replace']:
            existing = set(DictionaryEntry.objects.values_list('key', flat=True))

        objs = [
            DictionaryEntry(word=w, key=k, meaning=m)
            for w, k, m in clean if k not in existing
        ]
        DictionaryEntry.objects.bulk_create(objs, batch_size=2000, ignore_conflicts=True)
        self.stdout.write(self.style.SUCCESS(
            f'적재 {len(objs)}개 (사전 총 {DictionaryEntry.objects.count()}개)'))
