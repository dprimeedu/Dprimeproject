"""단어훈련 채점 로직.

영→한 주관식 테스트: 영어 단어를 보고 한글 뜻을 입력 → 채점.
뜻은 보통 쉼표로 여러 개(예: '값이 비싼, 귀중한')이고 괄호 보충(예: '평형 (균형)')도 있어,
그 중 하나만 맞아도 정답으로 인정한다.
"""
import random
import re

# 뜻 구분자: 쉼표·슬래시·세미콜론·가운뎃점 + ' 또는 '
_SPLIT_RE = re.compile(r'[,/;·∙・]|\s또는\s')
_PAREN_RE = re.compile(r'\(([^)]*)\)')
_STRIP_RE = re.compile(r'[\s()\[\]{}~\-–—.,/·∙・;:!?\'"`]')


def _norm(text):
    """채점용 정규화: 소문자 + 공백·구두점·괄호 제거."""
    if not text:
        return ''
    return _STRIP_RE.sub('', str(text).strip().lower())


def meaning_variants(meaning):
    """정답 뜻 문자열에서 인정 가능한 정규화 답안 집합을 만든다.

    '평형 (균형)'  → {'평형', '균형', '평형균형'}
    '값이 비싼, 귀중한' → {'값이비싼', '귀중한'}
    """
    out = set()
    for part in _SPLIT_RE.split(meaning or ''):
        part = part.strip()
        if not part:
            continue
        # 괄호 안 내용도 각각 독립 답안으로 인정
        for inner in _PAREN_RE.findall(part):
            for x in _SPLIT_RE.split(inner):
                n = _norm(x)
                if n:
                    out.add(n)
        # 괄호 제거 버전
        without = _norm(_PAREN_RE.sub('', part))
        if without:
            out.add(without)
        # 괄호 포함(괄호 기호만 제거) 버전
        full = _norm(part)
        if full:
            out.add(full)
    return out


def grade_meaning(student_input, correct_meaning):
    """학생이 입력한 한글 뜻이 정답 뜻의 한 갈래와 일치하면 True."""
    si = _norm(student_input)
    if not si:
        return False
    return si in meaning_variants(correct_meaning)


# ─────────────────────────────────────────────
# 시험 문제 추출
# ─────────────────────────────────────────────

def _wrong_rate_map(word_ids):
    """단어별 과거 오답률 (= 난이도 proxy). 시도 기록이 있는 단어만 반환."""
    from django.db.models import Count, Q
    from .models import VocabAttempt
    rows = (
        VocabAttempt.objects.filter(word_id__in=word_ids)
        .values('word_id')
        .annotate(total=Count('id'), wrong=Count('id', filter=Q(is_correct=False)))
    )
    return {r['word_id']: (r['wrong'] / r['total']) for r in rows if r['total']}


def select_test_words(student, unit, start_index, end_index, count=40, star_ratio=0.7):
    """범위(start~end 번호)에서 시험 문제용 단어를 골라 반환 (셔플된 VocabWord 리스트).

    규칙: 별표(모르는 단어) 70% 랜덤 + 비별표 30% 난이도 어려운 순.
    난이도 = 과거 오답률(기록 있을 때) → 없으면 단어 길이. 한쪽이 모자라면 다른 쪽에서 채움.
    """
    from .models import StudentWordStar

    words = list(
        unit.words.filter(index__gte=start_index, index__lte=end_index).order_by('index')
    )
    if not words:
        return []
    if len(words) <= count:
        random.shuffle(words)
        return words

    word_ids = [w.id for w in words]
    starred = set(
        StudentWordStar.objects
        .filter(student=student, word_id__in=word_ids)
        .values_list('word_id', flat=True)
    )
    starred_words = [w for w in words if w.id in starred]
    non_starred = [w for w in words if w.id not in starred]

    n_star = min(len(starred_words), round(count * star_ratio))
    n_non = count - n_star
    if n_non > len(non_starred):  # 비별표 부족 → 별표에서 더
        n_star = min(len(starred_words), n_star + (n_non - len(non_starred)))
        n_non = count - n_star

    # 별표: 랜덤
    random.shuffle(starred_words)
    pick = starred_words[:n_star]

    # 비별표: 난이도 어려운 순 (오답률 → 길이)
    diff = _wrong_rate_map([w.id for w in non_starred])
    non_sorted = sorted(
        non_starred,
        key=lambda w: (diff.get(w.id, 0.0), len(w.word or '')),
        reverse=True,
    )
    pick += non_sorted[:n_non]

    # 그래도 모자라면 남은 단어로 채움
    if len(pick) < count:
        chosen = {w.id for w in pick}
        rest = [w for w in words if w.id not in chosen]
        random.shuffle(rest)
        pick += rest[: count - len(pick)]

    random.shuffle(pick)
    return pick[:count]


def sync_pairs_to_dictionary(pairs):
    """[(영어, 한글)] 목록 중 통합 사전(DictionaryEntry)에 없는 단어만 추가.

    내신단어 등 단어 import 시 호출 → 통합 사전이 자동으로 커진다.
    기존 사전 항목(전체 단어장 모음 등)은 덮어쓰지 않고 '없는 것만' 추가.
    반환: 새로 추가한 개수.
    """
    from .models import DictionaryEntry

    seen = set()
    cleaned = []  # [(word, key, meaning)]
    for word, meaning in pairs:
        word = (str(word) if word is not None else '').strip()[:255]
        meaning = (str(meaning) if meaning is not None else '').strip()
        key = word.lower()
        if not (word and meaning and key) or key in seen:
            continue
        seen.add(key)
        cleaned.append((word, key, meaning))
    if not cleaned:
        return 0

    keys = [k for _, k, _ in cleaned]
    existing = set(
        DictionaryEntry.objects.filter(key__in=keys).values_list('key', flat=True)
    )
    objs = [
        DictionaryEntry(word=w, key=k, meaning=m)
        for w, k, m in cleaned if k not in existing
    ]
    if objs:
        DictionaryEntry.objects.bulk_create(objs, batch_size=2000, ignore_conflicts=True)
    return len(objs)


# ─────────────────────────────────────────────
# 단어장 xlsx 파싱 (영어/한글 컬럼 자동감지) — 사전·단어장 import 공용
# ─────────────────────────────────────────────
import re as _re

_HANGUL_RE = _re.compile(r'[가-힣]')
_ENG_WORD_RE = _re.compile(r"^[A-Za-z][A-Za-z'.\-/() ]{0,39}$")


def _cell_is_word(v):
    """영어 단어(짧은 형태)인지 — 예문/문장 컬럼 배제."""
    if v is None:
        return False
    s = str(v).strip()
    if not s or _HANGUL_RE.search(s) or not _ENG_WORD_RE.match(s):
        return False
    return len(s.split()) <= 4


def _cell_has_kor(v):
    return v is not None and bool(_HANGUL_RE.search(str(v)))


def extract_word_pairs(path, max_cols=12):
    """단어장 xlsx 첫 시트에서 (영어, 한글) 목록 추출. 영어/한글 컬럼 자동감지."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    ncol = min(max_cols, max((len(r) for r in rows[:50]), default=0))
    if not ncol:
        return []
    eng_score = [0] * ncol
    kor_score = [0] * ncol
    for r in rows[:80]:
        for ci in range(min(ncol, len(r))):
            if _cell_is_word(r[ci]):
                eng_score[ci] += 1
            if _cell_has_kor(r[ci]):
                kor_score[ci] += 1
    eng_col = max(range(ncol), key=lambda c: eng_score[c])
    kor_col = max(range(ncol), key=lambda c: kor_score[c])
    if eng_score[eng_col] < 3 or kor_score[kor_col] < 3:
        return []
    pairs = []
    for r in rows:
        if eng_col >= len(r) or kor_col >= len(r):
            continue
        w, m = r[eng_col], r[kor_col]
        if not _cell_is_word(w) or not _cell_has_kor(m):
            continue
        pairs.append((str(w).strip(), str(m).strip()))
    return pairs


# ─────────────────────────────────────────────
# 교재 단어장 배정 → 100단어 퀴즈렛 범위 자동 생성
# ─────────────────────────────────────────────
QUIZLET_SOURCE = '퀴즈렛'
QUIZLET_SIZE = 100


def _norm_word(s):
    """단어 채점용 정규화 — 공백 제거 + 소문자. (구글시트 C==H 정확일치와 동치)"""
    return (s or '').replace(' ', '').lower()


def grade_word(student_input, answer):
    """영영 시험 채점 — 학생이 고른 단어 == 정답 단어(정규화 후). 정답 비면 None."""
    if not (answer or '').strip():
        return None
    return _norm_word(student_input) == _norm_word(answer)


def split_range_into_chunks(start, end, size=QUIZLET_SIZE):
    """[start,end] 정수 범위를 100-경계에 맞춰 분할.

    101~300 → [(101,200),(201,300)]; 1~100 → [(1,100)]; 150~300 → [(150,200),(201,300)].
    내신단어TEST 범위를 퀴즈렛(1-100, 101-200…)과 같은 100단위 시험으로 쪼개는 데 사용.
    """
    start, end = int(start), int(end)
    if end < start:
        return [(start, end)]
    chunks = []
    s = start
    while s <= end:
        boundary = ((s + size - 1) // size) * size   # s 이상에서 가장 가까운 size 배수
        e = min(boundary, end)
        chunks.append((s, e))
        s = e + 1
    return chunks


def ensure_quizlet_ranges(student, unit, size=QUIZLET_SIZE, assigned_by=None, active_start=None):
    """교재 단어장을 학생에게 배정할 때 100단어 단위 퀴즈렛 범위 생성.

    이미 활성 퀴즈렛 범위가 있으면 스킵(중복 방지). 반환: 생성한 세트 수.

    active_start:
      · None  → 기존 동작(모든 청크 활성). 관리자 배정 UI 등 순차 아닌 경로 호환.
      · 정수  → **순차 모드**: active_start 가 포함된 청크 1개만 is_active=True,
               그 앞 청크는 '이미 한 것'(비활성), 뒤 청크는 대기(비활성). 서버가 리뷰 시 다음으로 진행.
    """
    from django.db.models import Count, Max
    from .models import VocabWord, VocabRangeTest

    agg = VocabWord.objects.filter(unit=unit).aggregate(mx=Max('index'), c=Count('id'))
    total = agg['mx'] or agg['c'] or 0
    if not total:
        return 0
    if VocabRangeTest.objects.filter(
        student=student, unit=unit, source_label__startswith=QUIZLET_SOURCE, is_active=True,
    ).exists():
        return 0

    # 청크 목록 (start, end) 순서대로
    chunks = []
    start = 1
    while start <= total:
        end = min(start + size - 1, total)
        chunks.append((start, end))
        start = end + 1

    sequential = active_start is not None
    active_idx = 0
    if sequential:
        try:
            a = int(active_start)
        except (TypeError, ValueError):
            a = 1
        for i, (cs, ce) in enumerate(chunks):
            if cs <= a <= ce:
                active_idx = i
                break

    rows = []
    for i, (cs, ce) in enumerate(chunks):
        rows.append(VocabRangeTest(
            student=student, unit=unit, start_index=cs, end_index=ce,
            source_label=QUIZLET_SOURCE, question_count=min(40, ce - cs + 1),
            time_limit_seconds=1200, pass_threshold=90, assigned_by=assigned_by,
            sort_order=i,
            is_active=(True if not sequential else (i == active_idx)),
        ))
    VocabRangeTest.objects.bulk_create(rows)
    return len(rows)


def activate_next_range(student, unit, source_label):
    """순차 진행 — 현재 활성 범위를 비활성화하고 '다음 미완료 청크'를 활성화.

    선생님 리뷰(is_reviewed=True) 시 호출. 다음 = start_index 오름차순 중 아직
    완료(finished+reviewed 세션)되지 않은 첫 청크. 없으면(끝까지 다 함) None.
    반환: 새로 활성화된 VocabRangeTest 또는 None.
    """
    from .models import VocabRangeTest

    current = (VocabRangeTest.objects
               .filter(student=student, unit=unit, source_label=source_label, is_active=True)
               .order_by('start_index').first())
    if current is None:
        return None
    current.is_active = False
    current.save(update_fields=['is_active'])

    nxt = None
    for rt in (VocabRangeTest.objects
               .filter(student=student, unit=unit, source_label=source_label,
                       is_active=False, start_index__gt=current.start_index)
               .order_by('start_index')):
        if not rt.sessions.filter(finished_at__isnull=False, is_reviewed=True).exists():
            nxt = rt
            break
    if nxt is not None:
        nxt.is_active = True
        nxt.save(update_fields=['is_active'])
    return nxt


def remove_quizlet_ranges(student, unit):
    """교재 단어장 배정 해제 시 그 단원의 퀴즈렛 범위 삭제. 반환: 삭제 수."""
    from .models import VocabRangeTest
    deleted, _ = VocabRangeTest.objects.filter(
        student=student, unit=unit, source_label__startswith=QUIZLET_SOURCE,
    ).delete()
    return deleted
